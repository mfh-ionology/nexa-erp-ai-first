#!/usr/bin/env python3
"""
Generate Nexa ERP Traceability Workbook
Comprehensive requirements → workflows → status rules → testing → architecture mapping
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

wb = openpyxl.Workbook()

# ─── Style Definitions ────────────────────────────────────────────────
HEADER_FONT = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
SUBHEADER_FILL = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
SUBHEADER_FONT = Font(name='Calibri', bold=True, size=11)
BODY_FONT = Font(name='Calibri', size=10)
WRAP = Alignment(wrap_text=True, vertical='top')
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def style_header(ws, row=1, max_col=None):
    if max_col is None:
        max_col = ws.max_column
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
        cell.border = THIN_BORDER

def style_body(ws, start_row=2, max_col=None):
    if max_col is None:
        max_col = ws.max_column
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, max_col=max_col):
        for cell in row:
            cell.font = BODY_FONT
            cell.alignment = WRAP
            cell.border = THIN_BORDER

def auto_width(ws, max_col=None, max_width=60):
    if max_col is None:
        max_col = ws.max_column
    for col in range(1, max_col + 1):
        max_len = 0
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 50), min_col=col, max_col=col):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)), max_width))
        ws.column_dimensions[get_column_letter(col)].width = max(max_len + 2, 12)

def add_section_row(ws, row, text, max_col):
    ws.cell(row=row, column=1, value=text)
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.border = THIN_BORDER

# ═══════════════════════════════════════════════════════════════════════
# SHEET 1: Module Features (All FRs)
# ═══════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = 'Module Features'
ws1.sheet_properties.tabColor = '2F5496'

headers = ['FR ID', 'Module', 'Category', 'Feature / Requirement', 'Architectural Location', 'Priority', 'Notes']
for col, h in enumerate(headers, 1):
    ws1.cell(row=1, column=col, value=h)

frs = [
    # AI Interaction
    ('FR1', 'AI', 'AI Interaction', 'Users can issue natural language commands to create, query, and manage business records across all modules', 'api/src/ai/orchestrator.ts', 'P0'),
    ('FR2', 'AI', 'AI Interaction', 'The system can pre-fill record fields using contextual knowledge (customer history, default terms, recent patterns)', 'api/src/ai/context-engine.ts', 'P0'),
    ('FR3', 'AI', 'AI Interaction', 'Users can receive a personalised daily briefing based on their role and system usage patterns', 'api/src/workers/briefing.worker.ts', 'P0'),
    ('FR4', 'AI', 'AI Interaction', 'Users can ask business questions in natural language and receive accurate answers with supporting data', 'api/src/ai/orchestrator.ts', 'P0'),
    ('FR5', 'AI', 'AI Interaction', 'The system can recommend actions to users with explanation and one-tap approval path', 'api/src/ai/orchestrator.ts', 'P0'),
    ('FR6', 'AI', 'AI Interaction', 'Users can approve, modify, or reject any AI-generated record or recommendation before it takes effect', 'api/src/ai/guardrails/', 'P0'),
    ('FR7', 'AI', 'AI Interaction', 'The system can maintain conversational context across a session to handle multi-step operations', 'api/src/ai/context-engine.ts + Redis', 'P0'),
    ('FR8', 'AI', 'AI Interaction', 'Users can fall back to traditional form-based interfaces for any operation at any time', 'web/src/ (dual interface pattern)', 'P0'),
    ('FR9', 'AI', 'AI Interaction', 'The system can log all AI actions, suggestions, and user responses for audit and learning purposes', 'AiConversation + AiMessage models', 'P0'),
    ('FR10', 'AI', 'AI Interaction', 'The system can display confidence scoring for AI-generated records and recommendations', 'api/src/ai/orchestrator.ts', 'P0'),
    # Finance/GL
    ('FR11', 'Finance', 'Finance/GL', 'Administrators can configure a chart of accounts based on UK GAAP (FRS 102) templates', 'modules/finance/ + seed data', 'P0'),
    ('FR12', 'Finance', 'Finance/GL', 'Users can create, edit, and post journal entries with double-entry enforcement', 'modules/finance/ + DB trigger', 'P0'),
    ('FR13', 'Finance', 'Finance/GL', 'Users can generate trial balance, profit & loss, and balance sheet reports for any date range', 'modules/reporting/', 'P0'),
    ('FR14', 'Finance', 'Finance/GL', 'Administrators can open and close financial periods with enforcement preventing posting to closed periods', 'modules/finance/ + DB trigger', 'P0'),
    ('FR15', 'Finance', 'Finance/GL', 'Users can manage multiple currencies with configurable exchange rates from external feeds', 'modules/system/ (Currency, ExchangeRate)', 'P0'),
    ('FR16', 'Finance', 'Finance/GL', 'Users can perform bank reconciliation matching bank statement lines to ledger transactions', 'modules/finance/bank-reconciliation/', 'P0'),
    ('FR17', 'Finance', 'Finance/GL', 'Users can import bank statements in OFX, CSV, and MT940 formats', 'integrations/banking/', 'P0'),
    ('FR18', 'Finance', 'Finance/GL', 'The system can automatically match bank transactions to open invoices and bills based on amount and reference', 'modules/finance/bank-reconciliation/', 'P0'),
    # AR
    ('FR19', 'AR', 'Accounts Receivable', 'Users can create and manage customer records with comprehensive fields (contacts, addresses, payment terms, VAT, credit limits, bank details)', 'modules/ar/customers/', 'P0'),
    ('FR20', 'AR', 'Accounts Receivable', 'Users can create, approve, and post sales invoices following the draft>approved>posted lifecycle', 'modules/ar/invoices/', 'P0'),
    ('FR21', 'AR', 'Accounts Receivable', 'Users can record customer payments and allocate them against specific invoices (full or partial)', 'modules/ar/payments/', 'P0'),
    ('FR22', 'AR', 'Accounts Receivable', 'Users can generate and send customer statements', 'modules/ar/ + modules/reporting/', 'P0'),
    ('FR23', 'AR', 'Accounts Receivable', 'Users can create credit notes linked to original invoices', 'modules/ar/invoices/', 'P0'),
    ('FR24', 'AR', 'Accounts Receivable', 'Users can view AR aging analysis by customer and date band', 'modules/reporting/', 'P0'),
    ('FR25', 'AR', 'Accounts Receivable', 'Users can manage multiple billing and shipping addresses per customer', 'modules/ar/customers/ (Address model)', 'P0'),
    # AP
    ('FR26', 'AP', 'Accounts Payable', 'Users can create and manage supplier records with comprehensive fields (contacts, addresses, payment terms, VAT, bank details)', 'modules/ap/suppliers/', 'P0'),
    ('FR27', 'AP', 'Accounts Payable', 'Users can create, approve, and post supplier bills following the draft>approved>posted lifecycle', 'modules/ap/bills/', 'P0'),
    ('FR28', 'AP', 'Accounts Payable', 'Users can record supplier payments and allocate them against specific bills', 'modules/ap/payments/', 'P0'),
    ('FR29', 'AP', 'Accounts Payable', 'Users can generate BACS payment files for bulk supplier payments', 'integrations/banking/bacs-generator.ts', 'P0'),
    ('FR30', 'AP', 'Accounts Payable', 'Users can view AP aging analysis by supplier and date band', 'modules/reporting/', 'P0'),
    ('FR31', 'AP', 'Accounts Payable', 'The system can match supplier bills to purchase orders (3-way matching with goods receipt)', 'modules/ap/bills/ (matching service)', 'P0'),
    ('FR32', 'AP', 'Accounts Payable', 'The system can ingest supplier bills via email/OCR and extract key fields for review', 'integrations/ocr/', 'P0'),
    # Sales
    ('FR33', 'Sales', 'Sales Management', 'Users can create sales quotes with line items, pricing, discounts, and VAT calculation', 'modules/sales/quotes/', 'P0'),
    ('FR34', 'Sales', 'Sales Management', 'Users can convert approved quotes to sales orders', 'modules/sales/ (quote-to-order service)', 'P0'),
    ('FR35', 'Sales', 'Sales Management', 'Users can manage sales orders through the full lifecycle (draft>confirmed>shipped>invoiced)', 'modules/sales/orders/', 'P0'),
    ('FR36', 'Sales', 'Sales Management', 'Users can create shipments/delivery notes against sales orders (full or partial)', 'modules/sales/shipments/', 'P0'),
    ('FR37', 'Sales', 'Sales Management', 'Users can convert sales orders to invoices (single or batch)', 'modules/sales/ (order-to-invoice service)', 'P0'),
    ('FR38', 'Sales', 'Sales Management', 'The system can check stock availability during order entry and alert on shortfalls', 'modules/inventory/ (stock check event)', 'P0'),
    ('FR39', 'Sales', 'Sales Management', 'Users can manage customer-specific pricing and discount rules', 'modules/sales/pricing/', 'P0'),
    ('FR40', 'Sales', 'Sales Management', 'Users can view sales pipeline with weighted values and activity tracking', 'modules/crm/pipeline/', 'P0'),
    # Purchasing
    ('FR41', 'Purchasing', 'Purchasing', 'Users can create purchase orders with line items, pricing, and VAT calculation', 'modules/purchasing/orders/', 'P0'),
    ('FR42', 'Purchasing', 'Purchasing', 'Users can manage PO approval workflows', 'modules/purchasing/orders/ (OKFlag)', 'P0'),
    ('FR43', 'Purchasing', 'Purchasing', 'Users can record goods receipt against purchase orders (full or partial)', 'modules/purchasing/goods-receipt/', 'P0'),
    ('FR44', 'Purchasing', 'Purchasing', 'The system can suggest reorder purchase orders based on stock levels and reorder points', 'modules/inventory/ (reorder service)', 'P0'),
    ('FR45', 'Purchasing', 'Purchasing', 'Users can track PO status through the lifecycle (draft>approved>ordered>received)', 'modules/purchasing/orders/', 'P0'),
    # Inventory
    ('FR46', 'Inventory', 'Inventory & Stock', 'Users can create and manage item records with typed relational fields (name, type, UoM, barcode, weight, cost price, sales price, VAT code, reorder point)', 'modules/inventory/items/', 'P0'),
    ('FR47', 'Inventory', 'Inventory & Stock', 'Users can manage item groups with default GL account mappings', 'modules/inventory/item-groups/', 'P0'),
    ('FR48', 'Inventory', 'Inventory & Stock', 'Users can record stock movements (receipt, issue, transfer, adjustment) with audit trail', 'modules/inventory/movements/', 'P0'),
    ('FR49', 'Inventory', 'Inventory & Stock', 'Users can manage multiple warehouse locations with bin-level tracking', 'modules/inventory/locations/', 'P0'),
    ('FR50', 'Inventory', 'Inventory & Stock', 'Users can perform stock takes with variance reporting', 'modules/inventory/stock-take/', 'P0'),
    ('FR51', 'Inventory', 'Inventory & Stock', 'Users can track items by serial number or batch number', 'modules/inventory/items/ (serial tracking)', 'P0'),
    ('FR52', 'Inventory', 'Inventory & Stock', 'Users can view real-time stock levels across all locations', 'modules/inventory/ + Redis cache', 'P0'),
    ('FR53', 'Inventory', 'Inventory & Stock', 'The system can alert users when items fall below reorder point', 'modules/inventory/ (reorder alert worker)', 'P0'),
    # CRM
    ('FR54', 'CRM', 'CRM', 'Users can create and manage contacts and accounts with activity history', 'modules/crm/contacts/', 'P0'),
    ('FR55', 'CRM', 'CRM', 'Users can log activities (calls, meetings, emails, notes) against contacts and accounts', 'modules/crm/activities/', 'P0'),
    ('FR56', 'CRM', 'CRM', 'Users can manage leads with status tracking and conversion to customers/quotes', 'modules/crm/leads/', 'P0'),
    ('FR57', 'CRM', 'CRM', 'Users can view pipeline reporting with stages, values, and probability weighting', 'modules/crm/pipeline/', 'P0'),
    ('FR58', 'CRM', 'CRM', 'Users can link CRM records to sales transactions (quotes, orders, invoices)', 'modules/crm/ (event bus linkage)', 'P0'),
    # HR
    ('FR59', 'HR', 'HR & Payroll', 'Users can create and manage employee records with typed fields required by UK employment law (NI number, tax code, employment type, start/end dates, department, job title, pay details)', 'modules/hr/employees/', 'P0'),
    ('FR60', 'HR', 'HR & Payroll', 'Users can manage employee onboarding with configurable checklist and self-service data collection', 'modules/hr/onboarding/', 'P0'),
    ('FR61', 'HR', 'HR & Payroll', 'Users can manage leave requests (request, approve, reject) with entitlement tracking', 'modules/hr/leave/', 'P0'),
    ('FR62', 'HR', 'HR & Payroll', 'Users can prepare and run monthly payroll with PAYE, NI, student loan, and pension calculations', 'modules/hr/payroll/ + integrations/payroll/', 'P0'),
    ('FR63', 'HR', 'HR & Payroll', 'The system can submit FPS and EPS to HMRC via RTI', 'integrations/hmrc/rti.adapter.ts', 'P0'),
    ('FR64', 'HR', 'HR & Payroll', 'Users can generate BACS payment files for payroll payments', 'integrations/banking/bacs-generator.ts', 'P0'),
    ('FR65', 'HR', 'HR & Payroll', 'The system can assess employee auto-enrolment pension eligibility and track opt-in/opt-out', 'modules/hr/payroll/ (pension service)', 'P0'),
    ('FR66', 'HR', 'HR & Payroll', 'Users can generate payslips, P45s, and P60s per HMRC specifications', 'modules/hr/payroll/ + modules/reporting/', 'P0'),
    ('FR67', 'HR', 'HR & Payroll', 'Users can manage statutory payments (SSP, SMP, SPP, ShPP, SAP)', 'modules/hr/payroll/ (statutory service)', 'P0'),
    # Manufacturing
    ('FR68', 'Manufacturing', 'Manufacturing', 'Users can create and manage Bills of Materials (BOM) with multi-level component structures', 'modules/manufacturing/bom/', 'P0'),
    ('FR69', 'Manufacturing', 'Manufacturing', 'Users can create and manage work orders with material requirements and routing operations', 'modules/manufacturing/work-orders/', 'P0'),
    ('FR70', 'Manufacturing', 'Manufacturing', 'Users can schedule production based on sales order priority and material availability', 'modules/manufacturing/scheduling/', 'P0'),
    ('FR71', 'Manufacturing', 'Manufacturing', 'Users can record material consumption against work orders', 'modules/manufacturing/work-orders/', 'P0'),
    ('FR72', 'Manufacturing', 'Manufacturing', 'Users can record finished goods receipt from completed work orders', 'modules/manufacturing/work-orders/', 'P0'),
    ('FR73', 'Manufacturing', 'Manufacturing', 'The system can check material availability before work order confirmation and alert on shortages', 'modules/manufacturing/ + modules/inventory/', 'P0'),
    # Reporting
    ('FR74', 'Reporting', 'Reporting', 'Users can generate standard financial reports (P&L, Balance Sheet, Trial Balance, Cash Flow)', 'modules/reporting/', 'P0'),
    ('FR75', 'Reporting', 'Reporting', 'Users can generate operational reports (AR/AP Aging, Stock Valuation, Bank Reconciliation)', 'modules/reporting/', 'P0'),
    ('FR76', 'Reporting', 'Reporting', 'Users can generate HR reports (Payslips, Employee List)', 'modules/reporting/', 'P0'),
    ('FR77', 'Reporting', 'Reporting', 'Users can generate VAT returns for HMRC MTD submission', 'integrations/hmrc/mtd-vat.adapter.ts', 'P0'),
    ('FR78', 'Reporting', 'Reporting', 'Users can export reports in PDF and CSV formats', 'modules/reporting/ (export service)', 'P0'),
    ('FR79', 'Reporting', 'Reporting', 'Users can ask ad-hoc business questions in natural language and receive data-backed answers', 'api/src/ai/ (NL query tool)', 'P0'),
    # Administration
    ('FR80', 'Admin', 'Administration', 'Administrators can create, edit, and deactivate user accounts with role assignment', 'core/rbac/', 'P0'),
    ('FR81', 'Admin', 'Administration', 'Administrators can assign roles (SUPER_ADMIN, ADMIN, MANAGER, STAFF, VIEWER) with module-level access gating', 'core/rbac/', 'P0'),
    ('FR82', 'Admin', 'Administration', 'Administrators can enable/disable modules per tenant configuration', 'core/tenant/ (module toggles)', 'P0'),
    ('FR83', 'Admin', 'Administration', 'Administrators can configure per-module settings (payment terms defaults, VAT schemes, number series, currency)', 'modules/system/ (SystemSetting)', 'P0'),
    ('FR84', 'Admin', 'Administration', 'Administrators can manage integration connections (bank feeds, HMRC, payroll API) with health monitoring', 'integrations/ + admin dashboard', 'P0'),
    ('FR85', 'Admin', 'Administration', 'Administrators can view audit logs of all system actions including AI-generated operations', 'core/audit/', 'P0'),
    ('FR86', 'Admin', 'Administration', 'Administrators can configure number series (prefix + sequential counter) per document type', 'modules/system/ (NumberSeries)', 'P0'),
    ('FR87', 'Admin', 'Administration', 'Administrators can import data (customers, suppliers, items, opening balances) from CSV files', 'modules/admin/ (CSV import worker)', 'P0'),
    ('FR88', 'Admin', 'Administration', 'Administrators can manage backup and restore operations', 'infrastructure (cloud ops)', 'P0'),
    # Compliance
    ('FR89', 'Compliance', 'Compliance & VAT', 'The system can calculate VAT at standard (20%), reduced (5%), zero (0%), exempt, and reverse charge rates', 'modules/system/ (VatCode) + calculation', 'P0'),
    ('FR90', 'Compliance', 'Compliance & VAT', 'Users can configure VAT schemes (Standard, Flat Rate, Cash Accounting)', 'modules/system/ (CompanyProfile.vatScheme)', 'P0'),
    ('FR91', 'Compliance', 'Compliance & VAT', 'The system can generate and submit VAT returns to HMRC via the MTD API', 'integrations/hmrc/mtd-vat.adapter.ts', 'P0'),
    ('FR92', 'Compliance', 'Compliance & VAT', 'The system can maintain immutable audit trails for all financial transactions', 'core/audit/ + DB rules', 'P0'),
    ('FR93', 'Compliance', 'Compliance & VAT', 'Administrators can manage GDPR compliance operations (data export, data deletion requests)', 'core/gdpr/', 'P0'),
    ('FR94', 'Compliance', 'Compliance & VAT', 'The system can enforce period locks preventing modification of closed financial period data', 'DB trigger: prevent_locked_period_modification', 'P0'),
    # CRM — Expanded
    ('FR95', 'CRM', 'CRM — Expanded', 'Users can create and manage marketing campaigns with recipient lists, media types, status tracking, and response analysis', 'modules/crm/campaigns/', 'P0'),
    ('FR96', 'CRM', 'CRM — Expanded', 'Users can manage sales opportunities with weighted pipeline stages, probability percentages, and expected revenue calculations', 'modules/crm/pipeline/', 'P0'),
    ('FR97', 'CRM', 'CRM — Expanded', 'Administrators can configure pipeline Kanban board stages at system level, and users can create personal pipeline views', 'modules/crm/pipeline/', 'P0'),
    ('FR98', 'CRM', 'CRM — Expanded', 'Administrators can define activity auto-creation rules that automatically generate activities for key CRM events', 'modules/crm/activities/', 'P0'),
    ('FR99', 'CRM', 'CRM — Expanded', 'Users can assign lead ratings (Cold, Warm, Hot) and manage lead lifecycle from initial contact through qualification to conversion or closure', 'modules/crm/leads/', 'P0'),
    ('FR100', 'CRM', 'CRM — Expanded', 'Administrators can configure CRM-specific activity types and activity groups for structured activity tracking', 'modules/crm/activities/', 'P0'),
    # HR & Payroll — Expanded
    ('FR101', 'HR', 'HR & Payroll — Expanded', 'Users can manage employment contracts through a full lifecycle (draft, approve, active, terminate) with immutable change history', 'modules/hr/contracts/', 'P0'),
    ('FR102', 'HR', 'HR & Payroll — Expanded', 'The system can track contract changes (salary adjustments, job title changes, department transfers) with full audit trail and effective dates', 'modules/hr/contracts/', 'P0'),
    ('FR103', 'HR', 'HR & Payroll — Expanded', 'Users can conduct performance appraisals using configurable factor and rating matrices with multi-level approval workflows', 'modules/hr/appraisals/', 'P0'),
    ('FR104', 'HR', 'HR & Payroll — Expanded', 'Users can manage employee skills and competencies with rating evaluations and point-in-time reporting', 'modules/hr/skills/', 'P0'),
    ('FR105', 'HR', 'HR & Payroll — Expanded', 'Users can create and manage training plans with scheduling, automatic double-booking detection, and auto-closure on completion', 'modules/hr/training/', 'P0'),
    ('FR106', 'HR', 'HR & Payroll — Expanded', 'Administrators can define job positions and organisational structure hierarchies for departmental reporting', 'modules/hr/positions/', 'P0'),
    ('FR107', 'HR', 'HR & Payroll — Expanded', 'Users can manage employee benefits on contracts with configurable benefit types, amounts, and payment frequencies', 'modules/hr/contracts/', 'P0'),
    ('FR108', 'HR', 'HR & Payroll — Expanded', 'Users can manage onboarding and offboarding checklists using configurable checkpoint templates with assignment and completion tracking', 'modules/hr/onboarding/', 'P0'),
    # Manufacturing — Expanded
    ('FR109', 'Manufacturing', 'Manufacturing — Expanded', 'Users can trigger recipe/BOM explosion across document types including quotes, sales orders, and invoices', 'modules/manufacturing/bom/', 'P0'),
    ('FR110', 'Manufacturing', 'Manufacturing — Expanded', 'Users can define and manage production shift schedules with worker assignments per shift', 'modules/manufacturing/scheduling/', 'P0'),
    ('FR111', 'Manufacturing', 'Manufacturing — Expanded', 'Users can register time worked by multiple workers per production operation with start/stop tracking', 'modules/manufacturing/work-orders/', 'P0'),
    ('FR112', 'Manufacturing', 'Manufacturing — Expanded', 'The system can post operation-level costs to GL accounts including work-in-progress (WIP) accounting entries', 'modules/manufacturing/work-orders/', 'P0'),
    ('FR113', 'Manufacturing', 'Manufacturing — Expanded', 'The system can run Material Requirements Planning (MRP) calculations based on demand, stock levels, lead times, and open orders', 'modules/manufacturing/mrp/', 'P0'),
    ('FR114', 'Manufacturing', 'Manufacturing — Expanded', 'Users can manage machine and work centre capacity with availability calendars and utilisation tracking', 'modules/manufacturing/scheduling/', 'P0'),
    ('FR115', 'Manufacturing', 'Manufacturing — Expanded', 'Users can perform quality inspections at the production operation level with pass/fail recording and defect tracking', 'modules/manufacturing/quality/', 'P0'),
    # POS (Phase 2)
    ('FR116', 'POS', 'POS', 'Users can open and close POS sessions with cash float entry, end-of-day cash counting, and Z-report generation', 'modules/pos/', 'P2'),
    ('FR117', 'POS', 'POS', 'Users can look up products by name, code, or barcode scan at the point of sale', 'modules/pos/', 'P2'),
    ('FR118', 'POS', 'POS', 'Users can process multiple payment methods per transaction including cash, card, voucher, and split payments', 'modules/pos/', 'P2'),
    ('FR119', 'POS', 'POS', 'Users can print receipts or send email receipts to customers from the POS terminal', 'modules/pos/', 'P2'),
    ('FR120', 'POS', 'POS', 'Administrators can configure POS-specific pricing rules and promotional discounts', 'modules/pos/', 'P2'),
    ('FR121', 'POS', 'POS', 'The system can operate in offline mode and automatically synchronise transactions when connectivity is restored', 'modules/pos/', 'P2'),
    ('FR122', 'POS', 'POS', 'Users can manage cash drawer operations with till reconciliation and variance reporting', 'modules/pos/', 'P2'),
    # Projects & Job Costing (Phase 2)
    ('FR123', 'Projects', 'Projects & Job Costing', 'Users can create and manage projects with budgets, phases, milestones, and status tracking', 'modules/projects/', 'P2'),
    ('FR124', 'Projects', 'Projects & Job Costing', 'Users can record time entries against projects with billable and non-billable classification', 'modules/projects/', 'P2'),
    ('FR125', 'Projects', 'Projects & Job Costing', 'Users can record expenses against projects with receipt attachment and approval workflow', 'modules/projects/', 'P2'),
    ('FR126', 'Projects', 'Projects & Job Costing', 'Users can view project budget versus actual reports with variance analysis by phase and cost category', 'modules/projects/', 'P2'),
    ('FR127', 'Projects', 'Projects & Job Costing', 'The system can resolve billing rates using a priority hierarchy: project-specific rate, customer rate, employee rate, then default rate', 'modules/projects/', 'P2'),
    ('FR128', 'Projects', 'Projects & Job Costing', 'The system can post job costs to GL accounts per project with cost centre tracking', 'modules/projects/', 'P2'),
    ('FR129', 'Projects', 'Projects & Job Costing', 'The system can calculate work-in-progress (WIP) values and support revenue recognition for long-running projects', 'modules/projects/', 'P2'),
    # Contracts & Agreements (Phase 2)
    ('FR130', 'Contracts', 'Contracts & Agreements', 'Users can create and manage contracts for rental, lease, and service agreements with defined terms and conditions', 'modules/contracts/', 'P2'),
    ('FR131', 'Contracts', 'Contracts & Agreements', 'The system can automatically generate recurring invoices from active contracts based on configured billing schedules', 'modules/contracts/', 'P2'),
    ('FR132', 'Contracts', 'Contracts & Agreements', 'Users can manage contract renewal and termination workflows with advance notification and approval steps', 'modules/contracts/', 'P2'),
    ('FR133', 'Contracts', 'Contracts & Agreements', 'Users can manage loan agreements with repayment schedule calculations supporting annuity, linear, and bullet repayment methods', 'modules/contracts/', 'P2'),
    ('FR134', 'Contracts', 'Contracts & Agreements', 'Users can configure contract-based pricing and payment terms that override default customer terms', 'modules/contracts/', 'P2'),
    # Warehouse Management (Phase 2)
    ('FR135', 'Warehouse', 'Warehouse Management', 'Users can define and manage warehouse positions and bin locations with capacity and zone classification', 'modules/warehouse/', 'P2'),
    ('FR136', 'Warehouse', 'Warehouse Management', 'Users can generate pick lists from sales orders and manage pick completion with quantity tracking', 'modules/warehouse/', 'P2'),
    ('FR137', 'Warehouse', 'Warehouse Management', 'Users can receive goods into specific warehouse positions with automatic or manual bin assignment', 'modules/warehouse/', 'P2'),
    ('FR138', 'Warehouse', 'Warehouse Management', 'Users can create internal transfer orders to move stock between warehouse positions', 'modules/warehouse/', 'P2'),
    ('FR139', 'Warehouse', 'Warehouse Management', 'Users can perform cycle counts by warehouse position with variance reporting and adjustment posting', 'modules/warehouse/', 'P2'),
    ('FR140', 'Warehouse', 'Warehouse Management', 'Users can manage packing and dispatch operations with shipment tracking and carrier assignment', 'modules/warehouse/', 'P2'),
    # Intercompany & Consolidation (Phase 3)
    ('FR141', 'Intercompany', 'Intercompany & Consolidation', 'The system can route intercompany transactions so that a PO in one company creates a corresponding SO in the counterpart', 'modules/intercompany/', 'P3'),
    ('FR142', 'Intercompany', 'Intercompany & Consolidation', 'The system can generate intercompany elimination journal entries for consolidated financial reporting', 'modules/intercompany/', 'P3'),
    ('FR143', 'Intercompany', 'Intercompany & Consolidation', 'Users can generate consolidated financial reports (P&L, Balance Sheet) across multiple companies', 'modules/intercompany/', 'P3'),
    ('FR144', 'Intercompany', 'Intercompany & Consolidation', 'The system can perform currency translation for foreign subsidiary consolidation using closing rate and average rate methods', 'modules/intercompany/', 'P3'),
    # Communications (Phase 3)
    ('FR145', 'Communications', 'Communications', 'Users can send and receive internal messages and notifications within the ERP system', 'modules/communications/', 'P3'),
    ('FR146', 'Communications', 'Communications', 'Users can send and receive emails from within the ERP with automatic linking to relevant business entities', 'modules/communications/', 'P3'),
    ('FR147', 'Communications', 'Communications', 'Users can view a chronological activity feed per entity showing all related interactions', 'modules/communications/', 'P3'),
    ('FR148', 'Communications', 'Communications', 'Users can attach documents to any business record with version tracking and access control', 'modules/communications/', 'P3'),
    # Service Orders (Phase 2)
    ('FR149', 'Service', 'Service Orders', 'Users can create and manage service orders with assignment to service personnel, status tracking, and completion recording', 'modules/service/', 'P2'),
    ('FR150', 'Service', 'Service Orders', 'Users can track service items (equipment or assets under service contract) with service history and warranty information', 'modules/service/', 'P2'),
    ('FR151', 'Service', 'Service Orders', 'Users can schedule field service visits with calendar integration and technician availability checking', 'modules/service/', 'P2'),
    ('FR152', 'Service', 'Service Orders', 'Users can convert completed service orders to invoices with automatic line item generation from service activities and parts used', 'modules/service/', 'P2'),
    # Reporting — AI-driven
    ('FR153', 'Reporting', 'Reporting', 'The system can generate AI-driven cash flow forecasts based on historical patterns, outstanding invoices, and known commitments', 'modules/reporting/ + AI', 'P1'),
    # Purchasing — Barcode scanning
    ('FR154', 'Purchasing', 'Purchasing', 'Users can use barcode scanning during goods receipt for rapid identification and quantity confirmation', 'modules/purchasing/ + mobile', 'P0'),
    # Compliance — Fraud Detection
    ('FR155', 'Compliance', 'Compliance & VAT', 'The system can detect and alert on duplicate payment attempts by matching supplier, amount, invoice reference, and date proximity', 'modules/ap/ + fraud detection', 'P0'),
    ('FR156', 'Compliance', 'Compliance & VAT', 'The system can flag suspicious transactions based on configurable rules (unusual amounts, out-of-pattern timing, new supplier with large first payment)', 'modules/ap/ + fraud detection', 'P0'),
    ('FR157', 'Compliance', 'Compliance & VAT', 'The system can generate a fraud risk summary report showing flagged transactions, duplicate payment attempts, and anomaly patterns', 'modules/reporting/', 'P0'),
    # Fixed Assets (Phase 2)
    ('FR158', 'Assets', 'Fixed Assets', 'Users can create and manage fixed asset records with acquisition date, cost, useful life, residual value, asset category, location, and responsible person', 'modules/assets/', 'P2'),
    ('FR159', 'Assets', 'Fixed Assets', 'The system can calculate depreciation using straight-line, reducing balance, and sum-of-digits methods per UK GAAP (FRS 102 Section 17)', 'modules/assets/', 'P2'),
    ('FR160', 'Assets', 'Fixed Assets', 'The system can automatically post monthly depreciation journal entries to the general ledger', 'modules/assets/', 'P2'),
    ('FR161', 'Assets', 'Fixed Assets', 'Users can record asset disposals with gain/loss calculation and GL posting', 'modules/assets/', 'P2'),
    ('FR162', 'Assets', 'Fixed Assets', 'Users can perform asset revaluations with revaluation reserve posting per FRS 102', 'modules/assets/', 'P2'),
    ('FR163', 'Assets', 'Fixed Assets', 'Users can generate a fixed asset register report showing cost, accumulated depreciation, and net book value per asset and category', 'modules/assets/', 'P2'),
    # Document Understanding (MVP)
    ('FR164', 'AI', 'Document Understanding', 'Users can upload, photograph, or forward financial documents for AI-powered data extraction', 'modules/document-understanding/', 'P0'),
    ('FR165', 'AI', 'Document Understanding', 'The system can extract structured fields from financial documents with field-level confidence scoring, achieving >85% extraction accuracy', 'modules/document-understanding/', 'P0'),
    ('FR166', 'AI', 'Document Understanding', 'The system can automatically match extracted document data to existing supplier records, purchase orders, and GL accounts', 'modules/document-understanding/', 'P0'),
    ('FR167', 'AI', 'Document Understanding', 'Users can review, correct, and approve AI-extracted document records before posting, with corrections feeding back to improve accuracy', 'modules/document-understanding/', 'P0'),
    ('FR168', 'AI', 'Document Understanding', 'The system can process documents in PDF, JPEG, PNG, and TIFF formats with automatic orientation correction and quality validation', 'modules/document-understanding/', 'P0'),
    # Document Knowledge Base (Phase 2)
    ('FR169', 'AI', 'Document Knowledge Base', 'Administrators can upload company documents for AI indexing and storage in a vector knowledge base', 'modules/document-understanding/', 'P2'),
    ('FR170', 'AI', 'Document Knowledge Base', 'Users can ask natural language questions about company policies, procedures, and employment terms, receiving accurate answers with source references', 'modules/document-understanding/', 'P2'),
    # Multi-Company Management
    ('FR171', 'System', 'Multi-Company', 'Administrators can create and manage multiple companies within a single tenant', 'core/company/ + Company model', 'P0'),
    ('FR172', 'System', 'Multi-Company', 'Users can switch between companies via the application header, with all subsequent queries scoped to the selected company', 'core/company/ + middleware', 'P0'),
    ('FR173', 'System', 'Multi-Company', 'Administrators can configure per-entity register sharing rules between companies', 'core/company/ + RegisterSharingRule', 'P0'),
    ('FR174', 'System', 'Multi-Company', 'The system must scope every database query by company, ensuring complete data isolation between companies', 'core/company/ + middleware', 'P0'),
    # Company-Specific RBAC
    ('FR175', 'System', 'Company RBAC', 'Administrators can assign users a global role that applies across all companies within the tenant', 'core/rbac/ + UserCompanyRole', 'P0'),
    ('FR176', 'System', 'Company RBAC', 'Administrators can assign per-company role overrides that take precedence over the global role for specific companies', 'core/rbac/ + UserCompanyRole', 'P0'),
    ('FR177', 'System', 'Company RBAC', 'The system must resolve user permissions by checking company-specific role first, then global role, with no access if neither exists', 'core/rbac/ + middleware', 'P0'),
    # i18n / Localization
    ('FR178', 'System', 'i18n / Localization', 'All user-facing text must use translation keys via a centralised translation system, with no hardcoded strings', 'packages/shared/i18n/', 'P0'),
    ('FR179', 'System', 'i18n / Localization', 'Administrators can set a default language per company, and users can select preferred language in profile settings', 'modules/system/ + Company.defaultLanguage', 'P0'),
    ('FR180', 'System', 'i18n / Localization', 'The system must format numbers, dates, and currency values according to the user locale using standard Intl APIs', 'packages/shared/i18n/', 'P0'),
    # Cross-Cutting Tasks
    ('FR181', 'System', 'Cross-Cutting Tasks', 'Users can create tasks from any business record with title, description, priority, and due date', 'core/tasks/ + Task model', 'P0'),
    ('FR182', 'System', 'Cross-Cutting Tasks', 'Users can assign tasks to one or more users, with task assignees receiving notifications on assignment', 'core/tasks/ + TaskAssignee', 'P0'),
    ('FR183', 'System', 'Cross-Cutting Tasks', 'Users can view and manage assigned tasks from a centralised task list with filtering', 'core/tasks/ + web/tasks/', 'P0'),
    # Notifications
    ('FR184', 'System', 'Notifications', 'The system can deliver real-time notifications via in-app (WebSocket), push (mobile), and email channels', 'core/notifications/', 'P0'),
    ('FR185', 'System', 'Notifications', 'Users can configure per-channel, per-event-type notification preferences', 'core/notifications/ + user prefs', 'P0'),
    ('FR186', 'System', 'Notifications', 'Users can view a notification centre showing all recent notifications with read/unread status', 'core/notifications/ + web/notifications/', 'P0'),
    # Email Integration
    ('FR187', 'System', 'Email Integration', 'The system can send business documents via email using configurable per-company SMTP settings', 'core/email/ + SMTP', 'P0'),
    ('FR188', 'System', 'Email Integration', 'Users can send emails from within the ERP directly from business records with document PDF attachment', 'core/email/ + document-templates', 'P0'),
    ('FR189', 'System', 'Email Integration', 'Administrators can configure email templates per document type with merge fields for dynamic content', 'core/email/ + templates', 'P0'),
    # Printer Management
    ('FR190', 'System', 'Printer Management', 'Users can configure print preferences per document type (auto-print on save, manual print, or download PDF only)', 'core/printing/ + user prefs', 'P0'),
    ('FR191', 'System', 'Printer Management', 'The system can generate PDF documents via the Document Templates engine and present them via browser Print API', 'core/printing/ + document-templates', 'P0'),
    ('FR192', 'System', 'Printer Management', 'Administrators can set default print preferences per document type at the company level', 'core/printing/ + Company settings', 'P0'),
]

for i, fr in enumerate(frs, 2):
    for j, val in enumerate(fr, 1):
        ws1.cell(row=i, column=j, value=val)

style_header(ws1, max_col=7)
style_body(ws1, max_col=7)
auto_width(ws1, max_col=7)

# ═══════════════════════════════════════════════════════════════════════
# SHEET 2: NFR Requirements
# ═══════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet('NFR Requirements')
ws2.sheet_properties.tabColor = '548235'

nfr_headers = ['NFR ID', 'Category', 'Requirement', 'Metric / Target', 'Architecture Enforcement', 'Notes']
for col, h in enumerate(nfr_headers, 1):
    ws2.cell(row=1, column=col, value=h)

nfrs = [
    ('NFR1', 'Performance', 'AI conversational responses within 3 seconds for 95th percentile', '< 3s (P95)', 'Streaming via WebSocket; Redis context cache', ''),
    ('NFR2', 'Performance', 'Traditional CRUD operations within 500ms for 95th percentile', '< 500ms (P95)', 'Fastify 70k req/s; Prisma query optimisation', ''),
    ('NFR3', 'Performance', 'Standard report generation within 5 seconds for up to 100K transactions', '< 5s', 'Redis report cache (15min TTL); aggregated queries', ''),
    ('NFR4', 'Performance', 'Bank feed processing within 15 minutes of feed availability', '< 15min', 'BullMQ worker; exponential backoff retry', ''),
    ('NFR5', 'Performance', 'Bulk operations (month-end, payroll for 250 employees) within 60 seconds', '< 60s', 'BullMQ worker; progress indication via WebSocket', ''),
    ('NFR6', 'Performance', 'Page load time under 2 seconds on 10Mbps+', '< 2s', 'Vite code splitting; lazy module loading', ''),
    ('NFR7', 'Performance', 'Support 50 concurrent users per tenant without degradation', '50 users/tenant', 'PgBouncer connection pooling; Redis sessions', ''),
    ('NFR8', 'Security', 'All data encrypted at rest (AES-256) and in transit (TLS 1.3)', 'AES-256 + TLS 1.3', 'PostgreSQL encryption; nginx TLS termination', ''),
    ('NFR9', 'Security', 'Database-per-tenant with zero cross-tenant access', 'Zero cross-tenant', 'Separate PostgreSQL databases; tenant middleware', ''),
    ('NFR10', 'Security', 'MFA support (TOTP at minimum)', 'TOTP MFA', 'RFC 6238 TOTP; required for ADMIN+', ''),
    ('NFR11', 'Security', 'Sessions expire after configurable inactivity (default 30 min)', '30min default', 'JWT access token (15min) + Redis refresh token', ''),
    ('NFR12', 'Security', 'All API endpoints authenticated and authorised', '100% coverage', 'Fastify onRequest hook; RBAC guard middleware', ''),
    ('NFR13', 'Security', 'Password storage using Argon2id', 'Argon2id', 'Argon2id (memory-hard, OWASP recommended)', ''),
    ('NFR14', 'Security', 'All financial modifications in immutable audit trail', '100% logging', 'AuditLog model + DB rules (no update/delete)', ''),
    ('NFR15', 'Security', 'Failed logins rate-limited (max 5 per 15 min)', '5/15min', 'Redis rate limiter; account lockout', ''),
    ('NFR16', 'Security', 'AI never auto-executes financial operations without user approval', '100% human approval', 'Guardrail chain (architecture 6.9)', 'IMMUTABLE RULE'),
    ('NFR17', 'Reliability', 'System uptime 99.9% (max 8.76h unplanned downtime/year)', '99.9%', 'Health endpoints; container orchestration', ''),
    ('NFR18', 'Reliability', 'Zero data loss for committed financial transactions (ACID)', '100% ACID', 'PostgreSQL ACID; Prisma transactions', ''),
    ('NFR19', 'Reliability', 'Automated daily backups with point-in-time recovery', 'Daily backup', 'pg_dump + WAL archiving', ''),
    ('NFR20', 'Reliability', 'Backup restoration within 1 hour for databases up to 50GB', '< 1h recovery', 'pg_restore; tested quarterly', ''),
    ('NFR21', 'Reliability', 'AI layer unavailable: all traditional form operations continue', '100% fallback', 'Dual interface pattern; no AI dependency', ''),
    ('NFR22', 'Reliability', 'External integration failures handled gracefully with retry and DLQ', 'Retry + DLQ', 'BullMQ; exponential backoff; dead-letter queue', ''),
    ('NFR23', 'Scalability', 'Support up to 1,000 tenants without architectural changes', '1,000 tenants', 'Database-per-tenant; LRU PrismaClient pool', ''),
    ('NFR24', 'Scalability', 'Each tenant handles up to 1M transactions per year', '1M txn/year', 'PostgreSQL indexing; partitioning if needed', ''),
    ('NFR25', 'Scalability', 'Per-tenant schema migrations without system-wide downtime', 'Zero-downtime', 'Per-tenant migration CLI; rolling updates', ''),
    ('NFR26', 'Scalability', 'New tenant provisioning within 60 seconds', '< 60s', 'Create DB + migrate + seed (automated)', ''),
    ('NFR27', 'Accessibility', 'WCAG 2.1 Level AA compliance for all form interfaces', 'WCAG 2.1 AA', 'Shadcn UI (accessible components)', ''),
    ('NFR28', 'Accessibility', 'All interactive elements keyboard-navigable', '100% keyboard nav', 'Shadcn built-in keyboard support', ''),
    ('NFR29', 'Accessibility', 'Colour contrast minimum 4.5:1 normal text, 3:1 large text', '4.5:1 / 3:1', 'Tailwind CSS colour system', ''),
    ('NFR30', 'Accessibility', 'AI chat supports screen reader compatibility', 'Screen reader', 'ARIA labels on chat components', ''),
    ('NFR31', 'Integration', 'All external APIs: retry with exponential backoff (max 3 retries)', '3 retries', 'Adapter pattern; BullMQ retry', ''),
    ('NFR32', 'Integration', 'HMRC MTD/RTI submissions within HMRC API timeout windows', 'HMRC compliant', 'integrations/hmrc/; timeout config', ''),
    ('NFR33', 'Integration', 'Bank feed sync: no data loss or duplicate transactions', 'Zero duplicates', 'Idempotency key on BankTransaction', ''),
    ('NFR34', 'Integration', 'Integration credentials stored encrypted, never in logs/API', '100% encrypted', 'AES-256 credential vault; log redaction', ''),
    ('NFR35', 'Integration', 'Integration health monitorable from admin dashboard with alerting', 'Dashboard alerts', 'Health check endpoints; Prometheus metrics', ''),
    ('NFR36', 'Data Integrity', 'Double-entry enforced at database level', '100% balanced', 'PostgreSQL AFTER INSERT trigger on journal_lines', ''),
    ('NFR37', 'Data Integrity', 'Period locks enforced at database level', '100% enforced', 'BEFORE INSERT/UPDATE/DELETE trigger', ''),
    ('NFR38', 'Data Integrity', 'All monetary values use fixed-point decimal', 'DECIMAL(19,4)', 'Prisma Decimal; no floating point', ''),
    ('NFR39', 'Data Integrity', 'Audit trail records append-only (no update or delete)', 'Immutable', 'DB rules: no_update_audit, no_delete_audit', ''),
    ('NFR40', 'Data Integrity', 'Data retention minimum 6 years for financial records per HMRC', '6-year minimum', 'Soft delete; no hard delete on financials', ''),
    ('NFR41', 'Maintainability', 'All code in TypeScript with strict mode', 'TS strict', 'tsconfig.json strict: true', ''),
    ('NFR42', 'Maintainability', 'All coding performed exclusively using Claude Opus 4.6', 'Opus 4.6 only', 'Development process constraint', ''),
    ('NFR43', 'Maintainability', 'Test coverage minimum 80% for business logic', '80% coverage', 'Vitest; co-located test files', ''),
    ('NFR44', 'Maintainability', 'Database schema changes via versioned migrations', 'Versioned', 'Prisma migrations; per-tenant', ''),
    ('NFR45', 'Maintainability', 'All API endpoints documented with OpenAPI/Swagger', 'OpenAPI', 'Fastify Swagger plugin', ''),
]

for i, nfr in enumerate(nfrs, 2):
    for j, val in enumerate(nfr, 1):
        ws2.cell(row=i, column=j, value=val)

style_header(ws2, max_col=6)
style_body(ws2, max_col=6)
auto_width(ws2, max_col=6)

# ═══════════════════════════════════════════════════════════════════════
# SHEET 3: Workflow Catalogue
# ═══════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet('Workflow Catalogue')
ws3.sheet_properties.tabColor = 'BF8F00'

wf_headers = ['Workflow ID', 'Workflow Name', 'Module', 'Goal / Outcome', 'Primary Actor', 'Supporting Roles', 'Trigger', 'Start Status', 'End Status', 'Key Entities', 'Related FRs', 'Notes']
for col, h in enumerate(wf_headers, 1):
    ws3.cell(row=1, column=col, value=h)

workflows = [
    ('WF-SYS-01', 'Tenant Provisioning & Onboarding', 'System', 'Create new SaaS tenant with isolated DB, defaults, ready for login', 'Super-Admin', 'System automation', 'New subscription / admin action', 'PENDING_SETUP', 'ACTIVE', 'Tenant, CompanyProfile, User, Currency, VatCode, PaymentTerms, NumberSeries', 'FR80-88, NFR26', '60s target provisioning'),
    ('WF-SYS-02', 'Chart of Accounts Setup', 'System/Finance', 'Configure GL accounts with control accounts, VAT linkages', 'Finance Admin', 'Accountant', 'Tenant setup wizard', 'DRAFT', 'ACTIVE', 'GlAccount, AccBlock, VatCode, SubLedgerControl', 'FR11', 'FRS 102 template seeded'),
    ('WF-FIN-01', 'Journal Entry Lifecycle', 'Finance', 'Create, approve, and post journal entries with double-entry validation', 'Accountant', 'Finance Manager', 'User creates journal entry', 'DRAFT', 'POSTED', 'JournalEntry, JournalLine, GlAccount, AuditLog', 'FR12, FR14', 'Double-entry trigger enforced at DB level'),
    ('WF-FIN-02', 'Financial Period Lifecycle', 'Finance', 'Control transaction window, prevent re-posting to closed periods', 'Finance Admin', 'Super-Admin', 'Admin opens/closes period', 'OPEN', 'LOCKED', 'FinancialPeriod, PeriodLockException, AuditLog', 'FR14, FR94', 'DB-level prevention trigger'),
    ('WF-FIN-03', 'Bank Reconciliation', 'Finance', 'Match bank transactions to GL entries, handle differences, post', 'Accountant', 'Finance Admin', 'Bank feed import / manual', 'UNMATCHED', 'RECONCILED', 'BankTransaction, BankAccount, BankMatch, JournalEntry, CustomerPayment, SupplierPayment', 'FR16-18', 'AI auto-match suggestions'),
    ('WF-AR-01', 'Sales Invoice Lifecycle', 'AR', 'Issue invoice, track payment, manage credit notes, enforce AR controls', 'Sales Admin', 'Finance Manager', 'User creates invoice / AI chat', 'DRAFT', 'PAID', 'CustomerInvoice, CustomerInvoiceLine, Customer, GlAccount, CustomerPayment, CreditNote, JournalEntry', 'FR19-25', '60+ validation rules (REQ-IV-*)'),
    ('WF-AR-02', 'Payment Allocation', 'AR', 'Receive and allocate customer payments to open invoices', 'Accountant', 'AR Clerk', 'Bank feed / manual receipt', 'UNALLOCATED', 'RECONCILED', 'CustomerPayment, CustomerInvoice, BankTransaction, JournalEntry', 'FR21', 'Partial & overpayment handling'),
    ('WF-AR-03', 'Credit Note Lifecycle', 'AR', 'Issue credit for returns/adjustments, apply to invoices or hold as credit', 'Sales Admin', 'Finance Manager', 'User creates credit note', 'DRAFT', 'APPLIED', 'CustomerInvoice (type=CREDIT_NOTE), Customer, JournalEntry', 'FR23', 'REQ-IV-070 to REQ-IV-079'),
    ('WF-AP-01', 'Purchase Bill Lifecycle', 'AP', 'Record supplier bills, match to POs, post to GL, manage payment schedule', 'AP Clerk', 'Finance Manager', 'Bill received (email/manual)', 'DRAFT', 'PAID', 'SupplierBill, SupplierBillLine, Supplier, PurchaseOrder, GoodsReceipt, JournalEntry', 'FR26-32', '3-way matching; OCR ingestion'),
    ('WF-AP-02', 'BACS Payment Run', 'AP/Banking', 'Select open AP invoices, generate BACS file, submit to bank', 'Finance Manager', 'AP Clerk, Treasurer', 'User starts payment run', 'DRAFT', 'CONFIRMED', 'PaymentRun, PaymentRunLine, SupplierBill, SupplierPayment, BankAccount, JournalEntry', 'FR28-29', 'UK BACS format; max 10M per file'),
    ('WF-SAL-01', 'Quote-to-Cash (Q2C)', 'Sales', 'Complete sales cycle from quote through delivery to invoice', 'Salesman', 'Warehouse, Billing', 'Customer inquiry / AI chat', 'QUOTE_DRAFT', 'INVOICED', 'SalesQuote, SalesOrder, SalesOrderLine, Shipment, CustomerInvoice, Stock', 'FR33-40', 'Credit limit check; stock reservation on approval'),
    ('WF-PUR-01', 'Purchase Order Lifecycle', 'Purchasing', 'Create PO, receive goods, match to invoice, manage supplier relationships', 'Buyer', 'Procurement Mgr, Warehouse', 'Reorder alert / manual', 'PO_DRAFT', 'CLOSED', 'PurchaseOrder, PurchaseOrderLine, Supplier, GoodsReceipt, SupplierBill, Stock', 'FR41-45', '3-way match: PO + GRN + Bill'),
    ('WF-INV-01', 'Stock Movement / Transfer', 'Inventory', 'Move stock between locations/warehouses, update FIFO, track history', 'Warehouse Manager', 'Stock Keeper', 'User initiates transfer', 'DRAFT', 'COMPLETED', 'StockMovement, StockMovementLine, Stock, Location, JournalEntry', 'FR48-49', 'Serial/batch validation; FIFO cost recalc'),
    ('WF-INV-02', 'Stock Take / Adjustment', 'Inventory', 'Physical count, reconcile to system stock, post adjustment GL entries', 'Stock Keeper', 'Warehouse Mgr, Finance', 'Period-end / ad-hoc', 'PENDING', 'POSTED', 'StockTake, StockTakeDetail, StockMovement, JournalEntry, InventoryItem', 'FR50, FR52', 'Cutoff period; variance reporting'),
    ('WF-CRM-01', 'Lead Lifecycle', 'CRM', 'Track sales leads from intake through conversion to customer', 'Salesman', 'Sales Manager', 'Web inquiry / referral / manual', 'NEW', 'CONVERTED', 'CrmLead, CrmContact, CrmAccount, CrmActivity, SalesQuote', 'FR54-58', 'Lead scoring; pipeline reporting'),
    ('WF-HR-01', 'Employee Lifecycle', 'HR', 'Track employee from hire through termination, manage leave, payroll', 'HR Manager', 'Payroll Admin', 'New hire / onboarding', 'ONBOARDING', 'TERMINATED', 'Employee, EmployeePayrollRecord, LeaveAllocation, Pension', 'FR59-61', 'UK employment law compliance'),
    ('WF-HR-02', 'Payroll Run', 'HR/Payroll', 'Calculate pay, manage deductions, submit RTI to HMRC, process payments', 'Payroll Admin', 'Finance Manager', 'Monthly schedule', 'PENDING', 'CONFIRMED', 'PayrollRun, PayrollRunLine, Employee, RTI, BankPayment, JournalEntry', 'FR62-67', 'PAYE, NI, pension, BACS'),
    ('WF-REP-01', 'Report Generation', 'Reporting', 'Generate financial/operational reports, export to PDF/CSV', 'Finance Manager', 'Any authorised user', 'On-demand / scheduled', 'PENDING', 'EXPORTED', 'Report definition, JournalEntry, GL aggregates', 'FR74-79', '10 P0 reports; AI NL queries'),
    ('WF-VAT-01', 'VAT Return Lifecycle', 'Compliance', 'Calculate VAT, generate return, submit to HMRC via MTD', 'Finance Admin', 'Accountant', 'Period-end', 'DRAFT', 'SUBMITTED', 'VatReturn, VatCode, CustomerInvoice, SupplierBill, HMRC MTD API', 'FR89-91, FR77', 'UK MTD compliance'),
    ('WF-AI-01', 'AI Chat Interaction', 'AI', 'Process user NL command through intent > plan > execute > confirm', 'User (any role)', 'AI Orchestrator', 'User types/speaks in chat', 'USER_INPUT', 'ACTION_TAKEN', 'AiConversation, AiMessage, AiAgent, AiPrompt, AiUsage', 'FR1-10', 'Guardrail chain for financial ops (NFR16)'),
    ('WF-AI-02', 'Daily Briefing', 'AI', 'Generate role-specific daily briefing with KPIs and action items', 'System (scheduled)', 'User (recipient)', 'BullMQ cron schedule', 'GENERATING', 'DELIVERED', 'AiConversation, AiMessage, Employee, Dashboard KPIs', 'FR3', 'Push notification (mobile)'),
    ('WF-MFG-01', 'Work Order Lifecycle', 'Manufacturing', 'Create work order, consume materials, record finished goods', 'Production Manager', 'Warehouse, Finance', 'Sales order / manual', 'DRAFT', 'COMPLETED', 'WorkOrder, Bom, BomLine, StockMovement, InventoryItem', 'FR68-73', 'Material availability check'),
]

for i, wf in enumerate(workflows, 2):
    for j, val in enumerate(wf, 1):
        ws3.cell(row=i, column=j, value=val)

style_header(ws3, max_col=12)
style_body(ws3, max_col=12)
auto_width(ws3, max_col=12)

# ═══════════════════════════════════════════════════════════════════════
# SHEET 4: WF-AR-01 Steps (Invoice Lifecycle - most complex)
# ═══════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet('WF-AR-01 Steps')
ws4.sheet_properties.tabColor = 'C00000'

step_headers = ['Step #', 'Phase', 'Actor', 'Page/Screen', 'Action', 'System Response', 'Required Data / Validation', 'Status Change', 'Actions/Tasks Created', 'Notifications', 'Audit Log', 'Notes / Edge Cases']
for col, h in enumerate(step_headers, 1):
    ws4.cell(row=1, column=col, value=h)

ar01_steps = [
    (1, 'Creation', 'Sales Admin / AI', 'Invoices > New Invoice', 'Create invoice; enter customer, date, line items, payment terms', 'Draft created; auto-save; number series assigned', 'REQ-IV-001: serial uniqueness\nREQ-IV-003: at least 1 row, Sum4 > 0\nREQ-IV-004: customer exists\nREQ-IV-005: customer not blocked\nREQ-IV-008: payment terms exist', 'None (DRAFT)', 'None', 'None', 'Create event: Invoice Created (Draft)', 'AI can draft from NL: "Invoice Acme Corp for 3x widgets at 50 each"'),
    (2, 'Line Entry', 'Sales Admin / AI', 'Invoice Detail', 'Add line items; set quantities, prices, VAT codes, revenue accounts', 'Lines saved; totals auto-calculated; VAT computed', 'REQ-IV-013: item exists\nREQ-IV-014: item not terminated\nREQ-IV-015: sales account exists\nREQ-IV-016: sales account not blocked\nREQ-IV-018: VAT code exists', 'None', 'None', 'None', 'Activity log: line updates', 'Multi-currency: exchange rate applied if foreign currency invoice'),
    (3, 'Validation', 'Sales Admin', 'Invoice Detail', 'Review invoice; check totals; verify customer details', 'Validation warnings shown (credit limit, GP margin)', 'REQ-IV-019: GP% check (if DisallowSaleBelowGP)\nREQ-IV-021: VAT number validation\nREQ-IV-026: date in valid period\nREQ-IV-029: negative totals blocked', 'None', 'None', 'None', 'Validation results logged', 'Credit limit check: balance + invoice < customer.creditLimit'),
    (4, 'Approval', 'Finance Manager', 'Invoice Detail (Admin)', 'Click Approve; system runs approval checks', 'OKFlag 0>1; GL posting triggered; invoice locked', 'Permission: InvOK required\nCredit limit passed\nAll validation rules passed\nDate not in locked period', 'DRAFT > APPROVED', 'GL Journal Entry created:\n- Dr: AR control (1100)\n- Cr: Revenue account\n- Cr: VAT account', 'Notify customer (if auto-send)\nNotify sales team', 'Status transition: who/when/from/to\nGL posting reference', 'FIFO update if stock items; Stock reservation released'),
    (5, 'Posting', 'System', 'Automatic', 'GL journal entry transitions to POSTED', 'Invoice immutable; appears in AR aging', 'GL journal balanced (double-entry trigger)', 'APPROVED > POSTED', 'None', 'None', 'GL posting confirmed', 'Automatic step after approval succeeds'),
    (6, 'Delivery', 'System / Admin', 'Email / PDF', 'Invoice PDF generated and emailed to customer', 'PDF with company branding; email sent via queue', 'Customer email address exists', 'None', 'Email queue task', 'Customer receives invoice email', 'Email delivery logged', 'PDF stored in document storage for future reference'),
    (7, 'Partial Payment', 'Accountant', 'Payments > Allocate', 'Record customer payment; allocate to invoice (partial)', 'Payment allocated; AR balance updated', 'Payment amount > 0\nCustomer matches\nInvoice is POSTED', 'POSTED > PARTIALLY_PAID', 'CustomerPayment created\nGL: Dr Bank, Cr AR control', 'Notify finance team', 'Payment allocation logged', 'Early payment discount applied if within discount window'),
    (8, 'Final Payment', 'Accountant', 'Payments > Allocate', 'Allocate remaining payment; invoice fully paid', 'Balance = 0; invoice removed from aging', 'Full amount allocated', 'PARTIALLY_PAID > PAID', 'GL posting for final payment', 'Optional: thank-you notification', 'Final payment logged', 'Overpayment creates customer credit memo'),
    (9, 'Void (alt)', 'Finance Manager', 'Invoice Detail', 'Void invoice if error discovered post-approval', 'Reverse GL entry created automatically; invoice marked VOIDED', 'Permission required\nReason mandatory\nDate not in locked period', 'POSTED > VOIDED', 'Reverse journal entry\nCredit note auto-created', 'Notify customer', 'Void reason + reversal logged', 'Cannot void if payments already allocated; must credit note instead'),
]

for i, step in enumerate(ar01_steps, 2):
    for j, val in enumerate(step, 1):
        ws4.cell(row=i, column=j, value=val)

style_header(ws4, max_col=12)
style_body(ws4, max_col=12)
auto_width(ws4, max_col=12)

# ═══════════════════════════════════════════════════════════════════════
# SHEET 5: WF-SAL-01 Steps (Quote-to-Cash)
# ═══════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet('WF-SAL-01 Steps')
ws5.sheet_properties.tabColor = 'ED7D31'

for col, h in enumerate(step_headers, 1):
    ws5.cell(row=1, column=col, value=h)

sal01_steps = [
    (1, 'Quote Creation', 'Salesman / AI', 'Sales > New Quote', 'Create quote for customer; add line items, pricing, discounts', 'Quote draft created; pricing calculated; VAT applied', 'Customer exists & not blocked\nItems exist & active\nPrices > 0\nValidity date set', 'None (QUOTE_DRAFT)', 'None', 'None', 'Quote created (Draft)', 'AI: "Quote Acme Corp 100x Product A at 50 each"'),
    (2, 'Quote Send', 'Salesman', 'Quote Detail', 'Send quote to customer (email PDF)', 'Quote PDF generated; email sent', 'Customer email exists\nAll lines valid', 'QUOTE_DRAFT > QUOTE_SUBMITTED', 'Email task queued', 'Customer receives quote PDF', 'Quote sent logged', 'Validity countdown starts'),
    (3, 'Convert to Order', 'Salesman', 'Quote Detail', 'Customer accepts; click Convert to Sales Order', 'Sales Order created from quote data; quote linked', 'Quote not expired\nCustomer still active', 'QUOTE > ORDER_DRAFT', 'SalesOrder created', 'Notify warehouse (pending)', 'Conversion logged', 'All quote lines copied to order lines'),
    (4, 'Order Approval', 'Sales Manager', 'Order Detail', 'Approve sales order; credit & stock checks run', 'OKFlag 0>1; stock reserved; planned payment created', 'REQ-OR-005: credit limit check\nREQ-OR-009: date in period\nREQ-OR-016: items valid\nREQ-OR-024: GP% check', 'ORDER_DRAFT > ORDER_APPROVED', 'Stock reservation created\nCRM follow-up activity\nPlanned payment record', 'Notify warehouse: ready to ship\nNotify customer: order confirmed', 'Order approval logged\nStock reservation logged', 'Backorder handling if stock insufficient'),
    (5, 'Pick & Pack', 'Warehouse Staff', 'Shipments > Pick List', 'Pick items from warehouse; scan barcodes; pack into boxes', 'Stock deducted (FIFO); shipment record updated', 'Items available in location\nSerial numbers if tracked', 'None (Shipment: PENDING > PICKED > PACKED)', 'None', 'None', 'Pick/pack activity logged', 'Barcode scanning via mobile app'),
    (6, 'Ship', 'Warehouse Manager', 'Shipments > Ship', 'Mark shipment as shipped; enter tracking number', 'Delivery note PDF generated; tracking info saved', 'All items packed\nCarrier selected', 'ORDER_APPROVED > PARTIALLY_SHIPPED or SHIPPED', 'Delivery note PDF', 'Customer: "Your order is on the way"\nSalesman notified', 'Shipment logged', 'Partial shipment creates backorder for remaining'),
    (7, 'Delivery', 'Customer / Carrier', 'External', 'Customer receives goods; carrier confirms delivery', 'Delivery confirmed; POD recorded', 'Tracking confirms delivery', 'Shipment: SHIPPED > DELIVERED', 'None', 'Salesman: delivery confirmed', 'Delivery logged', 'Signature capture if applicable'),
    (8, 'Invoice Generation', 'Billing / System', 'Invoices', 'Create invoice from shipped order (manual or auto)', 'CustomerInvoice created with SO reference; all lines linked', 'All shipped lines included\nPricing matches order', 'ORDER_SHIPPED > ORDER_INVOICED', 'Invoice created (DRAFT)\nSee WF-AR-01 for invoice lifecycle', 'None (invoice workflow takes over)', 'Invoice creation logged', 'Batch invoicing supported for multiple SOs'),
]

for i, step in enumerate(sal01_steps, 2):
    for j, val in enumerate(step, 1):
        ws5.cell(row=i, column=j, value=val)

style_header(ws5, max_col=12)
style_body(ws5, max_col=12)
auto_width(ws5, max_col=12)

# ═══════════════════════════════════════════════════════════════════════
# SHEET 6: Status Rules - Invoices
# ═══════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet('Status Rules - Invoices')
ws6.sheet_properties.tabColor = 'C00000'

sr_headers = ['From Status', 'To Status', 'Role Allowed', 'Required Fields / Evidence', 'System Actions', 'Notifications', 'Notes']
for col, h in enumerate(sr_headers, 1):
    ws6.cell(row=1, column=col, value=h)

invoice_rules = [
    ('DRAFT', 'APPROVED', 'Finance Manager (InvOK permission)', 'All required fields; credit limit passed; GP% check; VAT number valid; date in open period', 'GL posting (Dr AR, Cr Revenue, Cr VAT); FIFO update; stock reservation release', 'Customer (auto-send PDF); sales team', 'OKFlag 0>1; RegDate=today; immutable after'),
    ('APPROVED', 'POSTED', 'System (automatic)', 'GL journal entry balanced (double-entry trigger)', 'Invoice appears in AR aging; customer AR balance updated', 'None', 'Automatic after GL posting succeeds'),
    ('POSTED', 'PARTIALLY_PAID', 'Accountant', 'Payment received; allocation to invoice', 'Payment GL posting (Dr Bank, Cr AR); AR balance updated', 'Finance team notification', 'Amount < invoice total'),
    ('PARTIALLY_PAID', 'PAID', 'Accountant', 'Final payment allocated; balance = 0', 'Invoice removed from aging; all allocations complete', 'Optional thank-you', 'Overpayment creates credit memo'),
    ('POSTED', 'VOIDED', 'Finance Manager', 'Reason mandatory; date not in locked period; no payments allocated', 'Reverse GL entry auto-created; credit note auto-generated', 'Customer notified', 'Cannot void if payments exist; use credit note instead'),
    ('DRAFT', 'CANCELLED', 'Sales Admin / Finance Manager', 'Reason optional', 'No GL impact (was never posted)', 'None', 'Soft delete; record preserved for audit'),
    ('APPROVED', 'DRAFT (Un-approve)', 'Finance Manager', 'Period still open; no payments received', 'Reverse GL posting; reverse stock effects; reverse FIFO', 'Notify sales team', 'REQ-IV: Un-OK reverses all cross-entity effects'),
]

for i, rule in enumerate(invoice_rules, 2):
    for j, val in enumerate(rule, 1):
        ws6.cell(row=i, column=j, value=val)

style_header(ws6, max_col=7)
style_body(ws6, max_col=7)
auto_width(ws6, max_col=7)

# ═══════════════════════════════════════════════════════════════════════
# SHEET 7: Status Rules - Sales Orders
# ═══════════════════════════════════════════════════════════════════════
ws7 = wb.create_sheet('Status Rules - Orders')
ws7.sheet_properties.tabColor = 'ED7D31'

for col, h in enumerate(sr_headers, 1):
    ws7.cell(row=1, column=col, value=h)

order_rules = [
    ('DRAFT', 'APPROVED', 'Sales Manager (OROK)', 'Credit limit check; all items valid; payment terms; delivery address; GP% check', 'Stock reserved; CRM activity created; planned payment; SMS (if configured)', 'Warehouse notified; customer confirmation', 'OKFlag 0>1; REQ-OR-050 to REQ-OR-054'),
    ('APPROVED', 'PARTIALLY_SHIPPED', 'Warehouse Manager', 'Shipment created; items picked/packed/shipped', 'Stock deducted (FIFO); shipment record; delivery note PDF', 'Customer: partial shipment notification', 'Backorder created for remaining items'),
    ('APPROVED', 'SHIPPED', 'Warehouse Manager', 'All lines shipped', 'All stock deducted; all shipments complete', 'Customer: order shipped notification', 'All order lines fulfilled'),
    ('SHIPPED', 'INVOICED', 'Billing / System', 'Invoice created from SO', 'CustomerInvoice created; links to SO', 'None (invoice workflow)', 'Auto-invoice option configurable'),
    ('INVOICED', 'CLOSED', 'System / Sales Manager', 'All invoices PAID', 'Order archived; no further changes', 'None', 'Auto-close when all invoices paid'),
    ('DRAFT', 'CANCELLED', 'Sales Admin', 'Reason required', 'No stock impact', 'Customer notification optional', 'Cannot cancel after any shipment'),
    ('APPROVED', 'CANCELLED', 'Sales Manager', 'No items shipped; no payments; reason required', 'Stock reservations released; planned payments deleted', 'Customer + warehouse notified', 'REQ-OR-060 to REQ-OR-066 delete guards'),
]

for i, rule in enumerate(order_rules, 2):
    for j, val in enumerate(rule, 1):
        ws7.cell(row=i, column=j, value=val)

style_header(ws7, max_col=7)
style_body(ws7, max_col=7)
auto_width(ws7, max_col=7)

# ═══════════════════════════════════════════════════════════════════════
# SHEET 8: Status Rules - Bills (AP)
# ═══════════════════════════════════════════════════════════════════════
ws8 = wb.create_sheet('Status Rules - Bills')
ws8.sheet_properties.tabColor = '7030A0'

for col, h in enumerate(sr_headers, 1):
    ws8.cell(row=1, column=col, value=h)

bill_rules = [
    ('DRAFT', 'MATCHED_TO_PO', 'AP Clerk', 'PO match found; 3-way match within tolerance', 'PO reference linked; GRN linked', 'None', 'AI suggests PO matches'),
    ('MATCHED_TO_PO', 'APPROVED', 'Finance Manager (PUOK)', 'All validations passed; cost variance within threshold', 'Stock update; cost price recalc; serial tracking; PO update; GL posting (Dr Cost, Cr AP)', 'AP team notification', 'REQ-PU-030 to REQ-PU-039'),
    ('DRAFT', 'APPROVED', 'Finance Manager (PUOK)', 'No PO match needed (non-PO bill); all required fields', 'GL posting (Dr Cost, Cr AP)', 'AP team notification', 'Direct bill without PO'),
    ('APPROVED', 'POSTED', 'System (automatic)', 'GL journal balanced', 'Bill in AP aging; supplier AP balance updated', 'None', 'Automatic after GL posting'),
    ('POSTED', 'PARTIALLY_PAID', 'Finance Manager', 'Payment run includes this bill (partial amount)', 'Payment GL (Dr AP, Cr Bank); AP balance updated', 'None', 'Part of WF-AP-02 payment run'),
    ('PARTIALLY_PAID', 'PAID', 'Finance Manager', 'Final payment; balance = 0', 'Bill removed from AP aging', 'None', 'Full payment via BACS run'),
    ('POSTED', 'VOIDED', 'Finance Manager', 'Reason required; no payments allocated; period open', 'Reverse GL; reverse stock effects', 'Supplier notification optional', 'REQ-PU-040 to REQ-PU-043 un-OK effects'),
    ('APPROVED', 'DRAFT (Un-approve)', 'Finance Manager', 'Period open; no payments', 'Reverse stock; reverse serial numbers; recalculate cost', 'None', 'Un-OK: REQ-PU-040 to REQ-PU-043'),
]

for i, rule in enumerate(bill_rules, 2):
    for j, val in enumerate(rule, 1):
        ws8.cell(row=i, column=j, value=val)

style_header(ws8, max_col=7)
style_body(ws8, max_col=7)
auto_width(ws8, max_col=7)

# ═══════════════════════════════════════════════════════════════════════
# SHEET 9: Status Rules - Employees
# ═══════════════════════════════════════════════════════════════════════
ws9 = wb.create_sheet('Status Rules - Employees')
ws9.sheet_properties.tabColor = '00B050'

for col, h in enumerate(sr_headers, 1):
    ws9.cell(row=1, column=col, value=h)

emp_rules = [
    ('ONBOARDING', 'ACTIVE', 'HR Manager', 'Onboarding checklist complete; all required fields', 'Leave allocations created; pension auto-enrolment assessed; payroll eligible', 'Employee welcome email; manager notified', 'System access granted on transition'),
    ('ACTIVE', 'ON_LEAVE', 'HR Manager / Employee', 'Leave type; dates; manager approval', 'Payroll adjusted (SSP/SMP if statutory)', 'Manager notified; team notified', 'Return date tracked; auto-return on date'),
    ('ON_LEAVE', 'ACTIVE', 'HR Manager / System', 'Return date reached or manual return', 'Payroll restored; leave balance updated', 'Manager notified', 'Auto-return on expected date'),
    ('ACTIVE', 'OFFBOARDING', 'HR Manager', 'Resignation/termination; notice date; final day set', 'Exit checklist created; final pay calculated; access revocation scheduled', 'Employee; manager; IT team', 'Payroll: accrued leave + final pay calculated'),
    ('OFFBOARDING', 'TERMINATED', 'HR Manager / System', 'Final day passed; exit complete', 'Final payroll run; RTI EPS to HMRC; P45 generated; pension transferred', 'Employee (P45); HMRC (EPS)', 'Record locked; cannot reactivate'),
    ('ACTIVE', 'TERMINATED', 'HR Manager (immediate)', 'Reason required; immediate termination', 'Same as OFFBOARDING > TERMINATED but immediate', 'Employee; HR; legal if required', 'Emergency termination path'),
]

for i, rule in enumerate(emp_rules, 2):
    for j, val in enumerate(rule, 1):
        ws9.cell(row=i, column=j, value=val)

style_header(ws9, max_col=7)
style_body(ws9, max_col=7)
auto_width(ws9, max_col=7)

# ═══════════════════════════════════════════════════════════════════════
# SHEET 10: Status Rules - Leads (CRM)
# ═══════════════════════════════════════════════════════════════════════
ws10 = wb.create_sheet('Status Rules - Leads')
ws10.sheet_properties.tabColor = '0070C0'

for col, h in enumerate(sr_headers, 1):
    ws10.cell(row=1, column=col, value=h)

lead_rules = [
    ('NEW', 'QUALIFIED', 'Salesman', 'Qualification activities completed; requirements gathered', 'Lead score updated; pipeline position updated', 'Sales manager notified', 'Score based on engagement signals'),
    ('QUALIFIED', 'CONVERTED', 'Salesman / Sales Manager', 'Customer/opportunity fields completed', 'CrmOpportunity or SalesQuote created; lead linked', 'Salesman notified; customer record created', 'Terminal state; REQ-CRM-2: cannot cancel after conversion'),
    ('NEW', 'LOST', 'Salesman', 'Reason required (budget cut, competitor, etc.)', 'Removed from pipeline forecast', 'Sales manager notified', 'Terminal state'),
    ('QUALIFIED', 'LOST', 'Salesman', 'Reason required', 'Removed from pipeline forecast', 'Sales manager notified', 'Terminal state'),
    ('NEW', 'CANCELLED', 'Salesman / Admin', 'Reason required (duplicate, test, invalid)', 'Soft-delete; REQ-CRM-1: no further transitions', 'None', 'REQ-CRM-3: idempotent cancel'),
    ('QUALIFIED', 'CANCELLED', 'Salesman / Admin', 'Reason required', 'Soft-delete', 'None', 'Terminal state'),
]

for i, rule in enumerate(lead_rules, 2):
    for j, val in enumerate(rule, 1):
        ws10.cell(row=i, column=j, value=val)

style_header(ws10, max_col=7)
style_body(ws10, max_col=7)
auto_width(ws10, max_col=7)

# ═══════════════════════════════════════════════════════════════════════
# SHEET 11: Traceability Matrix (FR → Arch → Workflow → Test)
# ═══════════════════════════════════════════════════════════════════════
ws11 = wb.create_sheet('Traceability Matrix')
ws11.sheet_properties.tabColor = '002060'

trace_headers = ['FR ID', 'Module', 'Feature Summary', 'Architecture Section', 'Mapped Workflow(s)', 'Status Rules Sheet', 'Legacy Reference', 'Test Coverage', 'Notes']
for col, h in enumerate(trace_headers, 1):
    ws11.cell(row=1, column=col, value=h)

trace_data = [
    ('FR1', 'AI', 'Natural language commands across modules', '6.1-6.3 AI Infrastructure', 'WF-AI-01', '', '', 'TODO', 'Tool-use pattern; intent recognition'),
    ('FR2', 'AI', 'Context-aware field pre-fill', '6.5 Context Engine', 'WF-AI-01', '', '', 'TODO', 'Redis-cached UserContext'),
    ('FR3', 'AI', 'Personalised daily briefing', '6.8 Daily Briefing Engine', 'WF-AI-02', '', '', 'TODO', 'BullMQ scheduled; role-specific'),
    ('FR4', 'AI', 'NL business questions with data answers', '6.1-6.3 AI Infrastructure', 'WF-AI-01', '', '', 'TODO', 'Query tool in AI toolkit'),
    ('FR5', 'AI', 'Action recommendations with one-tap approval', '6.3 Agent Registry', 'WF-AI-01', '', '', 'TODO', 'AI suggests; user approves'),
    ('FR6', 'AI', 'User approval before AI action takes effect', '6.9 Guardrail Architecture', 'WF-AI-01', '', '', 'TODO', 'NFR16: never auto-execute financials'),
    ('FR7', 'AI', 'Multi-step conversational context', '6.5 Context Engine', 'WF-AI-01', '', '', 'TODO', 'Redis session; AiConversation model'),
    ('FR8', 'AI', 'Fallback to traditional forms', '5. Frontend Architecture', 'WF-AI-01', '', '', 'TODO', 'Dual interface pattern (NFR21)'),
    ('FR9', 'AI', 'Log all AI actions for audit', '6.6 Observability & Audit', 'WF-AI-01', '', '', 'TODO', 'AiConversation + AiMessage models'),
    ('FR10', 'AI', 'Confidence scoring on AI output', '6.1-6.3 AI Infrastructure', 'WF-AI-01', '', '', 'TODO', 'Score displayed in UI'),
    ('FR11', 'Finance', 'FRS 102 chart of accounts templates', '2.10 System Module (seed data)', 'WF-SYS-02', '', 'AccVc (23 fields)', 'TODO', 'CoA template seeded per tenant'),
    ('FR12', 'Finance', 'Journal entry with double-entry', '2.4 Double-Entry Enforcement', 'WF-FIN-01', '', 'AccTransVc', 'TODO', 'DB trigger enforcement'),
    ('FR13', 'Finance', 'Trial balance, P&L, balance sheet', 'modules/reporting/', 'WF-REP-01', '', 'hal/Reports/', 'TODO', '10 P0 reports'),
    ('FR14', 'Finance', 'Period open/close with enforcement', '2.5 Period Locking', 'WF-FIN-02', '', 'PeriodBlock, DBLockBlock', 'TODO', 'DB trigger: prevent_locked_period_modification'),
    ('FR15', 'Finance', 'Multi-currency with exchange rates', '2.10 System Module (Currency, ExchangeRate)', 'WF-SYS-01', '', 'CurncyCodeVc (31 fields), ERVc (7 fields)', 'TODO', 'BoE API feed planned'),
    ('FR16', 'Finance', 'Bank reconciliation', 'modules/finance/bank-reconciliation/', 'WF-FIN-03', '', 'BankRecVc', 'TODO', 'AI auto-match suggestions'),
    ('FR17', 'Finance', 'Bank statement import (OFX, CSV, MT940)', 'integrations/banking/', 'WF-FIN-03', '', 'hal/Imports/Bank*', 'TODO', '3 format parsers'),
    ('FR18', 'Finance', 'Auto-match bank transactions', 'modules/finance/bank-reconciliation/', 'WF-FIN-03', '', '', 'TODO', 'Amount + reference + date matching'),
    ('FR19', 'AR', 'Customer records with comprehensive fields', 'modules/ar/customers/', 'WF-AR-01', '', 'CUVc (170+ fields)', 'TODO', 'Target ~80 fields (from 313 legacy)'),
    ('FR20', 'AR', 'Invoice lifecycle (draft>approved>posted)', 'modules/ar/invoices/', 'WF-AR-01', 'Status Rules - Invoices', 'IVVc (280+ fields), IVVcRecAction (60 rules)', 'TODO', 'REQ-IV-001 to REQ-IV-095'),
    ('FR21', 'AR', 'Payment allocation (full/partial)', 'modules/ar/payments/', 'WF-AR-02', 'Status Rules - Invoices', 'ARPayVc', 'TODO', 'Partial + overpayment handling'),
    ('FR22', 'AR', 'Customer statements', 'modules/ar/ + modules/reporting/', 'WF-REP-01', '', '', 'TODO', 'PDF/email delivery'),
    ('FR23', 'AR', 'Credit notes linked to invoices', 'modules/ar/invoices/', 'WF-AR-03', 'Status Rules - Invoices', 'IVVc (InvType=CREDIT_NOTE)', 'TODO', 'REQ-IV-070 to REQ-IV-079'),
    ('FR24', 'AR', 'AR aging analysis', 'modules/reporting/', 'WF-REP-01', '', '', 'TODO', 'By customer + date band'),
    ('FR25', 'AR', 'Multiple billing/shipping addresses', 'modules/ar/customers/ (Address model)', 'WF-AR-01', '', 'CUVc InvAddr0-4, DelAddr0-4', 'TODO', 'OQ-011: multi-address confirmed'),
    ('FR26', 'AP', 'Supplier records with comprehensive fields', 'modules/ap/suppliers/', 'WF-AP-01', '', 'VEVc (subset of CUVc)', 'TODO', 'Target ~50 fields'),
    ('FR27', 'AP', 'Bill lifecycle (draft>approved>posted)', 'modules/ap/bills/', 'WF-AP-01', 'Status Rules - Bills', 'PUVc, PUVcRecAction (28 rules)', 'TODO', 'REQ-PU-001 to REQ-PU-053'),
    ('FR28', 'AP', 'Supplier payment allocation', 'modules/ap/payments/', 'WF-AP-02', 'Status Rules - Bills', 'APPayVc', 'TODO', 'Full + partial payments'),
    ('FR29', 'AP', 'BACS payment files', 'integrations/banking/bacs-generator.ts', 'WF-AP-02', '', 'hal/Exports/BankEng*', 'TODO', 'UK BACS format; OQ-015'),
    ('FR30', 'AP', 'AP aging analysis', 'modules/reporting/', 'WF-REP-01', '', '', 'TODO', 'By supplier + date band'),
    ('FR31', 'AP', '3-way matching (PO>GRN>Bill)', 'modules/ap/bills/ (matching service)', 'WF-AP-01, WF-PUR-01', 'Status Rules - Bills', '', 'TODO', 'Price variance tolerance configurable'),
    ('FR32', 'AP', 'OCR bill ingestion', 'integrations/ocr/', 'WF-AP-01', '', '', 'TODO', 'Email attachment extraction'),
    ('FR33', 'Sales', 'Sales quotes with pricing/VAT', 'modules/sales/quotes/', 'WF-SAL-01', '', 'QTVc', 'TODO', 'AI drafting from NL'),
    ('FR34', 'Sales', 'Quote to order conversion', 'modules/sales/', 'WF-SAL-01', 'Status Rules - Orders', 'QTVc > ORVc', 'TODO', 'One-click conversion'),
    ('FR35', 'Sales', 'Sales order full lifecycle', 'modules/sales/orders/', 'WF-SAL-01', 'Status Rules - Orders', 'ORVc, ORVcRecAction (47 rules)', 'TODO', 'REQ-OR-001 to REQ-OR-066'),
    ('FR36', 'Sales', 'Shipments/delivery notes', 'modules/sales/shipments/', 'WF-SAL-01', 'Status Rules - Orders', 'DispatchVc', 'TODO', 'Partial shipment support'),
    ('FR37', 'Sales', 'Order to invoice conversion', 'modules/sales/', 'WF-SAL-01, WF-AR-01', '', '', 'TODO', 'Single or batch conversion'),
    ('FR38', 'Sales', 'Stock availability check on order', 'modules/inventory/ (event bus)', 'WF-SAL-01', '', '', 'TODO', 'Real-time check; shortfall alert'),
    ('FR39', 'Sales', 'Customer-specific pricing/discounts', 'modules/sales/pricing/', 'WF-SAL-01', '', 'PLDefVc', 'TODO', 'Price lists per customer/group'),
    ('FR40', 'Sales', 'Sales pipeline with weighted values', 'modules/crm/pipeline/', 'WF-CRM-01', '', '', 'TODO', 'CRM-to-Sales integration'),
    ('FR41', 'Purchasing', 'Purchase orders with pricing/VAT', 'modules/purchasing/orders/', 'WF-PUR-01', '', 'POVc, POVcRAction (30 rules)', 'TODO', 'REQ-PO-001 to REQ-PO-043'),
    ('FR42', 'Purchasing', 'PO approval workflow', 'modules/purchasing/orders/ (OKFlag)', 'WF-PUR-01', '', '', 'TODO', 'POOK permission required'),
    ('FR43', 'Purchasing', 'Goods receipt (full/partial)', 'modules/purchasing/goods-receipt/', 'WF-PUR-01', '', '', 'TODO', 'GRN creates stock movement'),
    ('FR44', 'Purchasing', 'Reorder suggestions from stock levels', 'modules/inventory/ (reorder service)', 'WF-PUR-01', '', '', 'TODO', 'AI suggests reorder POs'),
    ('FR45', 'Purchasing', 'PO status tracking', 'modules/purchasing/orders/', 'WF-PUR-01', '', '', 'TODO', 'Draft > Approved > Received'),
    ('FR46', 'Inventory', 'Item records with typed fields', 'modules/inventory/items/', 'WF-INV-01', '', 'INVc (211 fields), INVcRAction (8 rules)', 'TODO', 'Target ~50 typed fields (OQ-002)'),
    ('FR47', 'Inventory', 'Item groups with GL defaults', 'modules/inventory/item-groups/', '', '', 'ITVc (67 fields)', 'TODO', 'OQ-010: confirmed'),
    ('FR48', 'Inventory', 'Stock movements with audit trail', 'modules/inventory/movements/', 'WF-INV-01', '', 'StockMovVc, StockMovVcRAction (18 rules)', 'TODO', 'REQ-SM-001 to REQ-SM-038'),
    ('FR49', 'Inventory', 'Multi-warehouse with bin tracking', 'modules/inventory/locations/', 'WF-INV-01', '', 'LocationVc', 'TODO', 'System module: Location entity'),
    ('FR50', 'Inventory', 'Stock takes with variance reporting', 'modules/inventory/stock-take/', 'WF-INV-02', '', 'StockTakeVc', 'TODO', 'Cutoff period; recount support'),
    ('FR51', 'Inventory', 'Serial/batch number tracking', 'modules/inventory/items/ (serial)', 'WF-INV-01', '', 'StockReservVc', 'TODO', 'Qty=1 per serial; FIFO cost'),
    ('FR52', 'Inventory', 'Real-time stock levels', 'modules/inventory/ + Redis cache', 'WF-INV-01', '', '', 'TODO', 'Cached with event-driven invalidation'),
    ('FR53', 'Inventory', 'Reorder point alerts', 'modules/inventory/ (reorder alert worker)', '', '', '', 'TODO', 'BullMQ worker; push notification'),
    ('FR54', 'CRM', 'Contacts and accounts with activity', 'modules/crm/contacts/', 'WF-CRM-01', '', 'ContactVc (2 rules)', 'TODO', ''),
    ('FR55', 'CRM', 'Activity logging (calls, meetings, etc)', 'modules/crm/activities/', 'WF-CRM-01', '', 'ActVc (80+ fields)', 'TODO', 'Legacy Activity = CrmActivity'),
    ('FR56', 'CRM', 'Lead management with conversion', 'modules/crm/leads/', 'WF-CRM-01', 'Status Rules - Leads', 'LeadVc', 'TODO', 'REQ-CRM-1 to REQ-CRM-11'),
    ('FR57', 'CRM', 'Pipeline with probability weighting', 'modules/crm/pipeline/', 'WF-CRM-01', '', '', 'TODO', 'Weighted value forecasting'),
    ('FR58', 'CRM', 'CRM to sales transaction linking', 'modules/crm/ (event bus)', 'WF-CRM-01, WF-SAL-01', '', '', 'TODO', 'Event: lead.converted > quote'),
    ('FR59', 'HR', 'Employee records (UK employment law)', 'modules/hr/employees/', 'WF-HR-01', 'Status Rules - Employees', 'HRMPAVc (7 rules)', 'TODO', 'Target ~30 typed fields (OQ-004)'),
    ('FR60', 'HR', 'Employee onboarding with checklist', 'modules/hr/onboarding/', 'WF-HR-01', 'Status Rules - Employees', '', 'TODO', 'Self-service data collection'),
    ('FR61', 'HR', 'Leave management with entitlements', 'modules/hr/leave/', 'WF-HR-01', 'Status Rules - Employees', '', 'TODO', '25 days annual; SSP after 3 days'),
    ('FR62', 'HR', 'Monthly payroll (PAYE, NI, pension)', 'modules/hr/payroll/', 'WF-HR-02', '', 'HRMPayrollVc (5 rules)', 'TODO', 'UK PAYE + NI + pension calc'),
    ('FR63', 'HR', 'RTI submissions (FPS/EPS) to HMRC', 'integrations/hmrc/rti.adapter.ts', 'WF-HR-02', '', '', 'TODO', 'Government Gateway API'),
    ('FR64', 'HR', 'BACS payroll payments', 'integrations/banking/bacs-generator.ts', 'WF-HR-02', '', '', 'TODO', 'Same BACS engine as AP'),
    ('FR65', 'HR', 'Auto-enrolment pension assessment', 'modules/hr/payroll/ (pension)', 'WF-HR-01, WF-HR-02', '', '', 'TODO', 'Age 22+; earnings >= 12,570'),
    ('FR66', 'HR', 'Payslips, P45s, P60s', 'modules/hr/payroll/ + reporting/', 'WF-HR-02, WF-REP-01', '', '', 'TODO', 'HMRC specification format'),
    ('FR67', 'HR', 'Statutory payments (SSP, SMP, etc)', 'modules/hr/payroll/ (statutory)', 'WF-HR-02', '', '', 'TODO', 'UK statutory rates'),
    ('FR68', 'Manufacturing', 'Bills of Materials (multi-level)', 'modules/manufacturing/bom/', 'WF-MFG-01', '', 'ProdVc, ProdItemVc', 'TODO', 'Component structures'),
    ('FR69', 'Manufacturing', 'Work orders with materials/routing', 'modules/manufacturing/work-orders/', 'WF-MFG-01', '', 'ProdOrderVc, ProdOperationVc', 'TODO', ''),
    ('FR70', 'Manufacturing', 'Production scheduling', 'modules/manufacturing/scheduling/', 'WF-MFG-01', '', '', 'TODO', 'Priority + material availability'),
    ('FR71', 'Manufacturing', 'Material consumption recording', 'modules/manufacturing/work-orders/', 'WF-MFG-01', '', '', 'TODO', 'Stock deduction on consumption'),
    ('FR72', 'Manufacturing', 'Finished goods receipt', 'modules/manufacturing/work-orders/', 'WF-MFG-01', '', '', 'TODO', 'Stock addition on completion'),
    ('FR73', 'Manufacturing', 'Material availability check', 'modules/manufacturing/ + inventory/', 'WF-MFG-01', '', '', 'TODO', 'Pre-confirmation check'),
    ('FR74', 'Reporting', 'Standard financial reports (P&L, BS, TB, CF)', 'modules/reporting/', 'WF-REP-01', '', '', 'TODO', '4 core financial reports'),
    ('FR75', 'Reporting', 'Operational reports (aging, stock val, bank rec)', 'modules/reporting/', 'WF-REP-01', '', '', 'TODO', '3 operational reports'),
    ('FR76', 'Reporting', 'HR reports (payslips, employee list)', 'modules/reporting/', 'WF-REP-01', '', '', 'TODO', '2 HR reports'),
    ('FR77', 'Reporting', 'VAT returns for MTD', 'integrations/hmrc/mtd-vat.adapter.ts', 'WF-VAT-01', '', '', 'TODO', 'HMRC MTD API submission'),
    ('FR78', 'Reporting', 'PDF and CSV export', 'modules/reporting/ (export service)', 'WF-REP-01', '', '', 'TODO', 'PDF + CSV; Excel P1'),
    ('FR79', 'Reporting', 'NL ad-hoc queries via AI', 'api/src/ai/ (NL query tool)', 'WF-AI-01, WF-REP-01', '', '', 'TODO', 'AI translates NL to query'),
    ('FR80', 'Admin', 'User account CRUD with roles', 'core/rbac/', 'WF-SYS-01', '', 'UserVc (90+ fields)', 'TODO', '5 roles defined'),
    ('FR81', 'Admin', 'Role assignment (5 roles)', 'core/rbac/', 'WF-SYS-01', '', 'AccessVc', 'TODO', 'Module-level gating'),
    ('FR82', 'Admin', 'Module enable/disable per tenant', 'core/tenant/ (module toggles)', 'WF-SYS-01', '', '', 'TODO', 'Feature flags per tenant'),
    ('FR83', 'Admin', 'Per-module settings configuration', 'modules/system/ (SystemSetting)', 'WF-SYS-01', '', 'AccBlock, RoundBlock, etc.', 'TODO', 'Key-value settings store'),
    ('FR84', 'Admin', 'Integration management with health monitoring', 'integrations/ + admin dashboard', '', '', '', 'TODO', 'Bank, HMRC, payroll API'),
    ('FR85', 'Admin', 'Audit log viewing (including AI)', 'core/audit/', '', '', '', 'TODO', 'Immutable audit trail'),
    ('FR86', 'Admin', 'Number series configuration', 'modules/system/ (NumberSeries)', 'WF-SYS-01', '', 'SerNrTrackBlock, DNMVc', 'TODO', 'Already designed in arch 2.8'),
    ('FR87', 'Admin', 'CSV data import', 'modules/admin/ (CSV import worker)', '', '', '', 'TODO', 'Customers, suppliers, items, OB'),
    ('FR88', 'Admin', 'Backup and restore', 'infrastructure (cloud ops)', '', '', '', 'TODO', 'pg_dump; WAL archiving'),
    ('FR89', 'Compliance', 'VAT calculation (5 rates)', 'modules/system/ (VatCode)', 'WF-VAT-01', '', 'VATCodeBlock', 'TODO', 'Standard/Reduced/Zero/Exempt/RC'),
    ('FR90', 'Compliance', 'VAT scheme configuration', 'modules/system/ (Company)', 'WF-SYS-01', '', '', 'TODO', 'Standard, Flat Rate, Cash'),
    ('FR91', 'Compliance', 'VAT return MTD submission', 'integrations/hmrc/mtd-vat.adapter.ts', 'WF-VAT-01', '', 'VATDeclVc (5 rules)', 'TODO', 'HMRC Government Gateway'),
    ('FR92', 'Compliance', 'Immutable audit trail for financials', 'core/audit/ + DB rules', '', '', '', 'TODO', 'DB rules: no update/delete on audit_log'),
    ('FR93', 'Compliance', 'GDPR operations (export, delete)', 'core/gdpr/', '', '', '', 'TODO', 'Soft delete + data export'),
    ('FR94', 'Compliance', 'Period lock enforcement', 'DB trigger: prevent_locked_period_modification', 'WF-FIN-02', '', 'GlobalDBLockBlock, DBLockBlock', 'TODO', 'DB-level trigger on all financial tables'),
    # CRM — Expanded (FR95-FR100)
    ('FR95', 'CRM', 'Marketing campaigns with recipient lists', 'modules/crm/campaigns/', 'WF-CRM-01', '', '', 'TODO', ''),
    ('FR96', 'CRM', 'Sales opportunities with weighted pipeline', 'modules/crm/pipeline/', 'WF-CRM-01', '', '', 'TODO', ''),
    ('FR97', 'CRM', 'Pipeline Kanban stage configuration', 'modules/crm/pipeline/', 'WF-CRM-01', '', '', 'TODO', ''),
    ('FR98', 'CRM', 'Activity auto-creation rules', 'modules/crm/activities/', 'WF-CRM-01', '', '', 'TODO', ''),
    ('FR99', 'CRM', 'Lead ratings and lifecycle management', 'modules/crm/leads/', 'WF-CRM-01', 'Status Rules - Leads', '', 'TODO', ''),
    ('FR100', 'CRM', 'CRM activity types and groups', 'modules/crm/activities/', 'WF-CRM-01', '', '', 'TODO', ''),
    # HR — Expanded (FR101-FR108)
    ('FR101', 'HR', 'Employment contract lifecycle', 'modules/hr/contracts/', 'WF-HR-01', 'Status Rules - Employees', '', 'TODO', 'Immutable change history'),
    ('FR102', 'HR', 'Contract change tracking with audit trail', 'modules/hr/contracts/', 'WF-HR-01', '', '', 'TODO', 'Effective dates'),
    ('FR103', 'HR', 'Performance appraisals', 'modules/hr/appraisals/', 'WF-HR-01', '', '', 'TODO', 'Multi-level approval'),
    ('FR104', 'HR', 'Skills and competencies management', 'modules/hr/skills/', '', '', '', 'TODO', ''),
    ('FR105', 'HR', 'Training plans with scheduling', 'modules/hr/training/', '', '', '', 'TODO', 'Double-booking detection'),
    ('FR106', 'HR', 'Job positions and org structure', 'modules/hr/positions/', '', '', '', 'TODO', ''),
    ('FR107', 'HR', 'Employee benefits on contracts', 'modules/hr/contracts/', '', '', '', 'TODO', ''),
    ('FR108', 'HR', 'Onboarding/offboarding checklists', 'modules/hr/onboarding/', 'WF-HR-01', '', '', 'TODO', ''),
    # Manufacturing — Expanded (FR109-FR115)
    ('FR109', 'Manufacturing', 'BOM explosion across document types', 'modules/manufacturing/bom/', 'WF-MFG-01', '', '', 'TODO', ''),
    ('FR110', 'Manufacturing', 'Production shift schedules', 'modules/manufacturing/scheduling/', 'WF-MFG-01', '', '', 'TODO', ''),
    ('FR111', 'Manufacturing', 'Time tracking per operation', 'modules/manufacturing/work-orders/', 'WF-MFG-01', '', '', 'TODO', ''),
    ('FR112', 'Manufacturing', 'Operation-level GL posting with WIP', 'modules/manufacturing/work-orders/', 'WF-MFG-01', '', '', 'TODO', ''),
    ('FR113', 'Manufacturing', 'MRP calculations', 'modules/manufacturing/mrp/', 'WF-MFG-01', '', '', 'TODO', ''),
    ('FR114', 'Manufacturing', 'Machine and work centre capacity', 'modules/manufacturing/scheduling/', 'WF-MFG-01', '', '', 'TODO', ''),
    ('FR115', 'Manufacturing', 'Quality inspections', 'modules/manufacturing/quality/', 'WF-MFG-01', '', '', 'TODO', ''),
    # POS (FR116-FR122)
    ('FR116', 'POS', 'POS sessions with Z-reports', 'modules/pos/', '', '', '', 'TODO', 'Phase 2'),
    ('FR117', 'POS', 'Product lookup by name/code/barcode', 'modules/pos/', '', '', '', 'TODO', 'Phase 2'),
    ('FR118', 'POS', 'Multi-payment processing', 'modules/pos/', '', '', '', 'TODO', 'Phase 2'),
    ('FR119', 'POS', 'Receipt printing/emailing', 'modules/pos/', '', '', '', 'TODO', 'Phase 2'),
    ('FR120', 'POS', 'POS pricing rules and promotions', 'modules/pos/', '', '', '', 'TODO', 'Phase 2'),
    ('FR121', 'POS', 'Offline mode with sync', 'modules/pos/', '', '', '', 'TODO', 'Phase 2'),
    ('FR122', 'POS', 'Cash drawer operations', 'modules/pos/', '', '', '', 'TODO', 'Phase 2'),
    # Projects (FR123-FR129)
    ('FR123', 'Projects', 'Project CRUD with budgets and milestones', 'modules/projects/', '', '', '', 'TODO', 'Phase 2'),
    ('FR124', 'Projects', 'Time entries against projects', 'modules/projects/', '', '', '', 'TODO', 'Phase 2'),
    ('FR125', 'Projects', 'Expense recording with approval', 'modules/projects/', '', '', '', 'TODO', 'Phase 2'),
    ('FR126', 'Projects', 'Budget vs actual reports', 'modules/projects/', '', '', '', 'TODO', 'Phase 2'),
    ('FR127', 'Projects', 'Billing rate priority hierarchy', 'modules/projects/', '', '', '', 'TODO', 'Phase 2'),
    ('FR128', 'Projects', 'Job costs GL posting', 'modules/projects/', '', '', '', 'TODO', 'Phase 2'),
    ('FR129', 'Projects', 'WIP and revenue recognition', 'modules/projects/', '', '', '', 'TODO', 'Phase 2'),
    # Contracts (FR130-FR134)
    ('FR130', 'Contracts', 'Contract management for rental/lease/service', 'modules/contracts/', '', '', '', 'TODO', 'Phase 2'),
    ('FR131', 'Contracts', 'Recurring invoice generation', 'modules/contracts/', '', '', '', 'TODO', 'Phase 2'),
    ('FR132', 'Contracts', 'Contract renewal/termination workflows', 'modules/contracts/', '', '', '', 'TODO', 'Phase 2'),
    ('FR133', 'Contracts', 'Loan repayment schedule calculations', 'modules/contracts/', '', '', '', 'TODO', 'Phase 2'),
    ('FR134', 'Contracts', 'Contract-based pricing overrides', 'modules/contracts/', '', '', '', 'TODO', 'Phase 2'),
    # Warehouse (FR135-FR140)
    ('FR135', 'Warehouse', 'Bin locations with capacity/zone', 'modules/warehouse/', '', '', '', 'TODO', 'Phase 2'),
    ('FR136', 'Warehouse', 'Pick lists from sales orders', 'modules/warehouse/', '', '', '', 'TODO', 'Phase 2'),
    ('FR137', 'Warehouse', 'Goods receipt into positions', 'modules/warehouse/', '', '', '', 'TODO', 'Phase 2'),
    ('FR138', 'Warehouse', 'Internal transfer orders', 'modules/warehouse/', '', '', '', 'TODO', 'Phase 2'),
    ('FR139', 'Warehouse', 'Cycle counts with variance', 'modules/warehouse/', '', '', '', 'TODO', 'Phase 2'),
    ('FR140', 'Warehouse', 'Packing and dispatch', 'modules/warehouse/', '', '', '', 'TODO', 'Phase 2'),
    # Intercompany (FR141-FR144)
    ('FR141', 'Intercompany', 'Intercompany transaction routing', 'modules/intercompany/', '', '', '', 'TODO', 'Phase 3'),
    ('FR142', 'Intercompany', 'Elimination journal entries', 'modules/intercompany/', '', '', '', 'TODO', 'Phase 3'),
    ('FR143', 'Intercompany', 'Consolidated financial reports', 'modules/intercompany/', '', '', '', 'TODO', 'Phase 3'),
    ('FR144', 'Intercompany', 'Currency translation for consolidation', 'modules/intercompany/', '', '', '', 'TODO', 'Phase 3'),
    # Communications (FR145-FR148)
    ('FR145', 'Communications', 'Internal messages and notifications', 'modules/communications/', '', '', '', 'TODO', 'Phase 3'),
    ('FR146', 'Communications', 'Email send/receive with entity linking', 'modules/communications/', '', '', '', 'TODO', 'Phase 3'),
    ('FR147', 'Communications', 'Activity feed per entity', 'modules/communications/', '', '', '', 'TODO', 'Phase 3'),
    ('FR148', 'Communications', 'Document attachment with versioning', 'modules/communications/', '', '', '', 'TODO', 'Phase 3'),
    # Service Orders (FR149-FR152)
    ('FR149', 'Service', 'Service order management', 'modules/service/', '', '', '', 'TODO', 'Phase 2'),
    ('FR150', 'Service', 'Service item tracking', 'modules/service/', '', '', '', 'TODO', 'Phase 2'),
    ('FR151', 'Service', 'Field service scheduling', 'modules/service/', '', '', '', 'TODO', 'Phase 2'),
    ('FR152', 'Service', 'Service order to invoice conversion', 'modules/service/', '', '', '', 'TODO', 'Phase 2'),
    # Misc (FR153-FR157)
    ('FR153', 'Reporting', 'AI-driven cash flow forecasting', 'modules/reporting/ + AI', '', '', '', 'TODO', ''),
    ('FR154', 'Purchasing', 'Barcode scanning for goods receipt', 'modules/purchasing/ + mobile', 'WF-PUR-01', '', '', 'TODO', ''),
    ('FR155', 'Compliance', 'Duplicate payment detection', 'modules/ap/ + fraud detection', 'WF-AP-02', '', '', 'TODO', ''),
    ('FR156', 'Compliance', 'Suspicious transaction flagging', 'modules/ap/ + fraud detection', '', '', '', 'TODO', ''),
    ('FR157', 'Compliance', 'Fraud risk summary report', 'modules/reporting/', 'WF-REP-01', '', '', 'TODO', ''),
    # Fixed Assets (FR158-FR163)
    ('FR158', 'Assets', 'Fixed asset records', 'modules/assets/', '', '', '', 'TODO', 'Phase 2'),
    ('FR159', 'Assets', 'Depreciation calculation (3 methods)', 'modules/assets/', '', '', '', 'TODO', 'Phase 2'),
    ('FR160', 'Assets', 'Monthly depreciation GL posting', 'modules/assets/', '', '', '', 'TODO', 'Phase 2'),
    ('FR161', 'Assets', 'Asset disposal with gain/loss', 'modules/assets/', '', '', '', 'TODO', 'Phase 2'),
    ('FR162', 'Assets', 'Asset revaluation', 'modules/assets/', '', '', '', 'TODO', 'Phase 2'),
    ('FR163', 'Assets', 'Fixed asset register report', 'modules/assets/', '', '', '', 'TODO', 'Phase 2'),
    # Document Understanding (FR164-FR170)
    ('FR164', 'AI', 'Financial document upload/capture', 'modules/document-understanding/', 'WF-AI-01', '', '', 'TODO', ''),
    ('FR165', 'AI', 'AI field extraction with confidence scoring', 'modules/document-understanding/', 'WF-AI-01', '', '', 'TODO', '>85% accuracy target'),
    ('FR166', 'AI', 'Auto-match to suppliers/POs/GL accounts', 'modules/document-understanding/', 'WF-AI-01, WF-AP-01', '', '', 'TODO', ''),
    ('FR167', 'AI', 'Review and approve AI-extracted records', 'modules/document-understanding/', 'WF-AI-01', '', '', 'TODO', 'Corrections improve accuracy'),
    ('FR168', 'AI', 'Multi-format document processing', 'modules/document-understanding/', '', '', '', 'TODO', 'PDF, JPEG, PNG, TIFF'),
    ('FR169', 'AI', 'Company document AI indexing', 'modules/document-understanding/', '', '', '', 'TODO', 'Phase 2: vector knowledge base'),
    ('FR170', 'AI', 'NL questions about company documents', 'modules/document-understanding/', 'WF-AI-01', '', '', 'TODO', 'Phase 2: source citations'),
    # Multi-Company (FR171-FR174)
    ('FR171', 'System', 'Multi-company creation and management', 'core/company/ + Company model', 'WF-SYS-01', '', '', 'TODO', 'companyId on every table'),
    ('FR172', 'System', 'Company switching via header', 'core/company/ + middleware', '', '', '', 'TODO', 'Query scoping middleware'),
    ('FR173', 'System', 'Register sharing rules between companies', 'core/company/ + RegisterSharingRule', 'WF-SYS-01', '', '', 'TODO', 'NONE/ALL/SELECTED modes'),
    ('FR174', 'System', 'Company-scoped query enforcement', 'core/company/ + middleware', '', '', '', 'TODO', 'Every query includes companyId'),
    # Company RBAC (FR175-FR177)
    ('FR175', 'System', 'Global role assignment', 'core/rbac/ + UserCompanyRole', 'WF-SYS-01', '', '', 'TODO', 'companyId=NULL means global'),
    ('FR176', 'System', 'Per-company role overrides', 'core/rbac/ + UserCompanyRole', 'WF-SYS-01', '', '', 'TODO', 'Exception pattern'),
    ('FR177', 'System', 'RBAC resolution: specific > global > no access', 'core/rbac/ + middleware', '', '', '', 'TODO', ''),
    # i18n (FR178-FR180)
    ('FR178', 'System', 'Translation key system for all UI text', 'packages/shared/i18n/', '', '', '', 'TODO', 'No hardcoded strings'),
    ('FR179', 'System', 'Language preference per user/company', 'modules/system/ + Company.defaultLanguage', 'WF-SYS-01', '', '', 'TODO', 'Fallback: user > company > en'),
    ('FR180', 'System', 'Locale-based number/date/currency formatting', 'packages/shared/i18n/', '', '', '', 'TODO', 'Intl API'),
    # Tasks (FR181-FR183)
    ('FR181', 'System', 'Create tasks from any business record', 'core/tasks/ + Task model', '', '', '', 'TODO', 'Polymorphic entityType/entityId'),
    ('FR182', 'System', 'Multi-assignee task assignment', 'core/tasks/ + TaskAssignee', '', '', '', 'TODO', 'Notification on assignment'),
    ('FR183', 'System', 'Centralised task list with filtering', 'core/tasks/ + web/tasks/', '', '', '', 'TODO', ''),
    # Notifications (FR184-FR186)
    ('FR184', 'System', 'Multi-channel real-time notifications', 'core/notifications/', '', '', '', 'TODO', 'WebSocket + push + email'),
    ('FR185', 'System', 'Per-channel notification preferences', 'core/notifications/ + user prefs', 'WF-SYS-01', '', '', 'TODO', ''),
    ('FR186', 'System', 'Notification centre UI', 'core/notifications/ + web/', '', '', '', 'TODO', 'Read/unread status'),
    # Email Integration (FR187-FR189)
    ('FR187', 'System', 'SMTP outbound for business documents', 'core/email/ + SMTP', '', '', '', 'TODO', 'Per-company SMTP config'),
    ('FR188', 'System', 'Email from business records with PDF attachment', 'core/email/ + document-templates', '', '', '', 'TODO', ''),
    ('FR189', 'System', 'Email templates per document type', 'core/email/ + templates', 'WF-SYS-01', '', '', 'TODO', 'Merge fields'),
    # Printer Management (FR190-FR192)
    ('FR190', 'System', 'Print preferences per document type', 'core/printing/ + user prefs', 'WF-SYS-01', '', '', 'TODO', 'Auto-print/manual/download'),
    ('FR191', 'System', 'PDF generation via browser Print API', 'core/printing/ + document-templates', '', '', '', 'TODO', 'Puppeteer + Print API'),
    ('FR192', 'System', 'Company-level default print preferences', 'core/printing/ + Company settings', 'WF-SYS-01', '', '', 'TODO', 'User can override'),
]

for i, row in enumerate(trace_data, 2):
    for j, val in enumerate(row, 1):
        ws11.cell(row=i, column=j, value=val)

style_header(ws11, max_col=9)
style_body(ws11, max_col=9)
auto_width(ws11, max_col=9)

# ═══════════════════════════════════════════════════════════════════════
# SHEET 12: Testing Scripts (skeleton for each module)
# ═══════════════════════════════════════════════════════════════════════
ws12 = wb.create_sheet('Testing Scripts')
ws12.sheet_properties.tabColor = '00B050'

test_headers = ['FR Ref', 'Module', 'Test Name', 'Test Steps', 'Expected Results', 'Notes']
for col, h in enumerate(test_headers, 1):
    ws12.cell(row=1, column=col, value=h)

tests = [
    ('FR1', 'AI', 'NL command creates invoice', '1) Open AI chat\n2) Type "Create invoice for Acme Corp, 3x widgets at 50 each"\n3) Review AI-generated draft\n4) Confirm', 'Invoice draft created with correct customer, 3 lines, total 150, VAT calculated', 'AI guardrail: requires human confirm'),
    ('FR6', 'AI', 'AI financial action requires approval', '1) Ask AI to "post invoice INV-00042"\n2) AI presents confirmation dialog\n3) User approves', 'Invoice NOT posted until explicit user approval; NFR16 enforced', 'IMMUTABLE: AI never auto-executes financial ops'),
    ('FR8', 'AI', 'Traditional form fallback works when AI unavailable', '1) Disable AI service\n2) Navigate to Invoices > New\n3) Fill form manually\n4) Save', 'Invoice created via traditional form; no AI dependency', 'NFR21: dual interface pattern'),
    ('FR11', 'Finance', 'FRS 102 CoA template loads on tenant creation', '1) Provision new tenant\n2) Navigate to Chart of Accounts\n3) Verify accounts exist', 'FRS 102 accounts seeded (Assets 1000-1999, Liabilities 2000-2999, etc.)', 'Seed data in tenant provisioning'),
    ('FR12', 'Finance', 'Journal entry enforces double-entry balance', '1) Create journal entry\n2) Add Dr line: 1100 = 100.00\n3) Try to post without Cr line\n4) Add Cr line: 4000 = 100.00\n5) Post', 'Step 3: posting blocked (unbalanced); Step 5: posting succeeds', 'DB trigger enforcement (NFR36)'),
    ('FR14', 'Finance', 'Period lock prevents posting to closed period', '1) Close period Jan 2026\n2) Try to create journal entry dated 15/01/2026\n3) Try to post', 'Posting blocked with error: "Period is closed"\nDB trigger fires', 'NFR37: DB-level enforcement'),
    ('FR15', 'Finance', 'Currency and exchange rate management', '1) Add currency USD\n2) Add exchange rate GBP/USD = 1.27\n3) Create foreign currency invoice\n4) Verify conversion', 'Invoice shows both GBP and USD amounts; GL posted in base currency (GBP)', 'System module: Currency + ExchangeRate'),
    ('FR19', 'AR', 'Customer creation with all required fields', '1) Navigate to Customers > New\n2) Fill: name, email, VAT number, payment terms, address\n3) Save', 'Customer created with all fields persisted; appears in customer list', 'isActive=true by default'),
    ('FR20', 'AR', 'Invoice lifecycle (draft > approved > posted)', '1) Create invoice (DRAFT)\n2) Approve invoice\n3) Verify GL posting\n4) Verify AR aging', 'Draft editable; Approved triggers GL posting; Posted is immutable', 'See WF-AR-01; 60+ validation rules'),
    ('FR21', 'AR', 'Partial payment allocation', '1) Post invoice for 1000\n2) Record payment of 600\n3) Allocate to invoice\n4) Verify status', 'Invoice status = PARTIALLY_PAID; remaining balance = 400', 'GL: Dr Bank 600, Cr AR 600'),
    ('FR23', 'AR', 'Credit note against existing invoice', '1) Create credit note linked to posted invoice\n2) Verify credit validations (REQ-IV-070)\n3) Approve credit note', 'Credit note created; total <= original AR balance; reverse GL posting', 'REQ-IV-070 to REQ-IV-079'),
    ('FR27', 'AP', 'Bill lifecycle with 3-way matching', '1) Create PO\n2) Record goods receipt\n3) Create bill\n4) Match to PO + GRN\n5) Approve bill', 'Bill matched; stock updated; cost price recalculated; GL posted', 'REQ-PU-030 to REQ-PU-039'),
    ('FR29', 'AP', 'BACS payment file generation', '1) Create payment run\n2) Select open bills\n3) Approve\n4) Generate BACS file\n5) Download', 'BACS file in UK format; correct supplier bank details; amounts match', 'Max 10M per file; GBP only'),
    ('FR35', 'Sales', 'Quote to order to shipment to invoice', '1) Create quote\n2) Convert to order\n3) Approve order (credit check)\n4) Ship items\n5) Create invoice', 'Full Q2C lifecycle; stock reserved on order approval; deducted on ship', 'See WF-SAL-01'),
    ('FR38', 'Sales', 'Stock availability check on order entry', '1) Set item stock = 10\n2) Create order for 15 units\n3) Check alert', 'Shortfall alert: "Only 10 available, 5 short"', 'Real-time stock check'),
    ('FR46', 'Inventory', 'Item creation with typed fields', '1) Create item with: name, barcode, cost price, sales price, UoM, reorder point\n2) Save', 'All typed fields persisted; item searchable by barcode', 'OQ-002: minimum 15 typed fields'),
    ('FR48', 'Inventory', 'Stock movement with audit trail', '1) Create stock transfer: 50 units from WH-A to WH-B\n2) Approve\n3) Verify stock levels', 'WH-A stock = -50; WH-B stock = +50; FIFO cost recalculated; audit logged', 'REQ-SM-001 to REQ-SM-038'),
    ('FR50', 'Inventory', 'Stock take with variance posting', '1) Initiate stock take\n2) Count items (system: 100, actual: 95)\n3) Reconcile\n4) Post adjustment', 'Variance = -5; GL: Dr COGS 5, Cr Stock 5; stock corrected', 'Cutoff period enforcement'),
    ('FR56', 'CRM', 'Lead lifecycle with conversion', '1) Create lead (NEW)\n2) Qualify lead\n3) Convert to opportunity/quote\n4) Verify cannot cancel after conversion', 'Lead status transitions correctly; REQ-CRM-2 enforced', 'Terminal states: CONVERTED, LOST'),
    ('FR59', 'HR', 'Employee record with UK fields', '1) Create employee: name, NI number, tax code, salary, department\n2) Save\n3) Verify NI format', 'Employee created with typed fields; NI number validated (2L+6D+1L)', 'UK employment law compliance'),
    ('FR62', 'HR', 'Monthly payroll run', '1) Start payroll for March 2026\n2) Calculate (PAYE, NI, pension)\n3) Review\n4) Approve\n5) Generate BACS + RTI', 'Correct gross/tax/NI/net per employee; BACS file; RTI submission file', 'UK PAYE + NI rates; auto-enrolment'),
    ('FR86', 'Admin', 'Number series configuration', '1) Navigate to System > Number Series\n2) Set invoice prefix = "INV-"\n3) Create invoice\n4) Verify number', 'Invoice numbered INV-00001; next = INV-00002 (gap-free)', 'PostgreSQL function: next_number()'),
    ('FR89', 'Compliance', 'VAT calculation at all 5 rates', '1) Create invoices with each VAT code (S, R, Z, E, RC)\n2) Verify VAT amounts', 'Standard=20%, Reduced=5%, Zero=0%, Exempt=0%, Reverse Charge=0%', 'VatCode seed data'),
    ('FR92', 'Compliance', 'Audit trail immutability', '1) Create and approve invoice\n2) Try to UPDATE audit_log directly\n3) Try to DELETE audit_log', 'UPDATE silently ignored; DELETE blocked; audit records immutable', 'DB rules: no_update_audit, no_delete_audit'),
    ('FR94', 'Compliance', 'Period lock enforcement at DB level', '1) Lock period\n2) Try INSERT journal_entry with date in locked period\n3) Try UPDATE existing entry\n4) Try DELETE', 'All 3 operations blocked by DB trigger', 'prevent_locked_period_modification trigger'),
    ('NFR1', 'Performance', 'AI response time < 3 seconds (P95)', '1) Send 100 AI chat messages\n2) Measure response times\n3) Calculate P95', 'P95 response time < 3 seconds', 'Streaming via WebSocket'),
    ('NFR2', 'Performance', 'CRUD operations < 500ms (P95)', '1) Perform 1000 CRUD operations\n2) Measure response times\n3) Calculate P95', 'P95 response time < 500ms', 'Fastify + Prisma + PgBouncer'),
    ('NFR9', 'Security', 'Database-per-tenant isolation', '1) Create 2 tenants\n2) Attempt cross-tenant query from Tenant A to Tenant B DB', 'Access denied; zero data leakage; separate PostgreSQL databases', 'No tenant_id columns; connection routing'),
    ('NFR16', 'Security', 'AI cannot auto-execute financial operations', '1) Configure AI agent to create+post invoice\n2) Execute\n3) Verify human approval required', 'AI creates draft but CANNOT post without explicit user confirmation', 'IMMUTABLE NFR16 rule; guardrail chain'),
    ('NFR36', 'Data Integrity', 'Double-entry enforcement at DB level', '1) Try to INSERT unbalanced journal lines via raw SQL\n2) Observe result', 'Insert blocked by PostgreSQL trigger; error raised', 'AFTER INSERT trigger on journal_lines'),
]

for i, test in enumerate(tests, 2):
    for j, val in enumerate(test, 1):
        ws12.cell(row=i, column=j, value=val)

style_header(ws12, max_col=6)
style_body(ws12, max_col=6)
auto_width(ws12, max_col=6)

# ═══════════════════════════════════════════════════════════════════════
# SAVE WORKBOOK
# ═══════════════════════════════════════════════════════════════════════
output_path = '/Users/mfh/MFH_Docs/My_Projects/nexa-erp-ai-first/_bmad-output/planning-artifacts/Nexa-ERP-Traceability-Workbook-v1.xlsx'
wb.save(output_path)
print(f"Workbook saved to: {output_path}")
print(f"Sheets: {wb.sheetnames}")
print(f"Sheet 1 (Module Features): {ws1.max_row} rows")
print(f"Sheet 2 (NFR Requirements): {ws2.max_row} rows")
print(f"Sheet 3 (Workflow Catalogue): {ws3.max_row} rows")
print(f"Sheet 4 (WF-AR-01 Steps): {ws4.max_row} rows")
print(f"Sheet 5 (WF-SAL-01 Steps): {ws5.max_row} rows")
print(f"Sheet 6 (Status Rules - Invoices): {ws6.max_row} rows")
print(f"Sheet 7 (Status Rules - Orders): {ws7.max_row} rows")
print(f"Sheet 8 (Status Rules - Bills): {ws8.max_row} rows")
print(f"Sheet 9 (Status Rules - Employees): {ws9.max_row} rows")
print(f"Sheet 10 (Status Rules - Leads): {ws10.max_row} rows")
print(f"Sheet 11 (Traceability Matrix): {ws11.max_row} rows")
print(f"Sheet 12 (Testing Scripts): {ws12.max_row} rows")
